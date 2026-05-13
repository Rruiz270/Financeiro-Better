#!/usr/bin/env python3
"""
Generate receitas_customers.js by:
1. Parsing receitas_all_data.js (Vindi bills — monthly installments)
2. Reading the CORRECT spreadsheet (3,328 real sales/purchases)
3. Fetching Vindi customer details via API (email, CPF for Vindi-only)
4. Cross-referencing by email and name
5. Producing customerData with CORRECT classification:
   - compras = real purchases (spreadsheet entries)
   - vindi_bills = monthly installments (Vindi recurrence)
   - renovacao = true only if compras > 1 OR marked "Renovação" in spreadsheet
"""

import base64
import json
import os
import re
import time
import unicodedata
import urllib.request
from datetime import datetime

import openpyxl

# ── Config ────────────────────────────────────────────────────────────
SPREADSHEET = "/Users/Raphael/Downloads/Vendas para Emissão de NF (5).xlsx"
VINDI_DATA_JS = "/Users/Raphael/Financeiro-Better/receitas_all_data.js"
OUTPUT_JS = "/Users/Raphael/Financeiro-Better/receitas_customers.js"
VINDI_KEY = "pXgDGOG6I5xaYamYFgkjkx0vnqO65rksLWBaU3YIZQU"


# ── Helpers ───────────────────────────────────────────────────────────
def norm(s: str) -> str:
    """Normalize a string for matching: strip, uppercase, remove accents."""
    if not s:
        return ""
    s = str(s).strip().upper()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return s


def safe_date(val) -> str:
    """Convert openpyxl cell value to YYYY-MM-DD string safely."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    # Some cells have serial values that aren't valid dates
    try:
        s = str(val).strip()
        # If it looks like a date string already
        if re.match(r"\d{4}-\d{2}-\d{2}", s):
            return s[:10]
        return s
    except Exception:
        return ""


def safe_phone(val) -> str:
    """Convert phone cell to string, handling float values like 5511984780104.0."""
    if val is None:
        return ""
    try:
        s = str(val).strip()
        if s.endswith(".0"):
            s = s[:-2]
        # Remove non-numeric chars except +
        return s
    except Exception:
        return ""


def safe_float(val) -> float:
    """Safely convert a cell value to float."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def safe_str(val) -> str:
    """Safely convert a cell value to stripped string."""
    if val is None:
        return ""
    try:
        return str(val).strip()
    except Exception:
        return ""


# ══════════════════════════════════════════════════════════════════════
# STEP 1: Parse Vindi bills from receitas_all_data.js
# ══════════════════════════════════════════════════════════════════════
print("=" * 60)
print("STEP 1: Parsing Vindi bills from receitas_all_data.js")
print("=" * 60)

with open(VINDI_DATA_JS, "r") as f:
    raw = f.read()

m = re.search(r"var data = (\[.*\]);?\s*$", raw, re.DOTALL)
if not m:
    raise ValueError("Could not parse receitas_all_data.js")

bills_raw = json.loads(m.group(1))
print(f"  Loaded {len(bills_raw)} Vindi bills")

# Group by cliente
vindi_by_customer: dict[str, dict] = {}
for b in bills_raw:
    c = b["cliente"]
    if c not in vindi_by_customer:
        vindi_by_customer[c] = {
            "bills": [],
            "total": 0.0,
            "first": b["data"],
            "last": b["data"],
        }
    rec = vindi_by_customer[c]
    rec["bills"].append({
        "data": b["data"],
        "valor": b["valor"],
        "situacao": b["situacao"],
        "mes": b["mes"],
        "categoria": b["categoria"],
    })
    rec["total"] += b["valor"]
    if b["data"] < rec["first"]:
        rec["first"] = b["data"]
    if b["data"] > rec["last"]:
        rec["last"] = b["data"]

print(f"  Unique Vindi customers: {len(vindi_by_customer)}")


# ══════════════════════════════════════════════════════════════════════
# STEP 2: Read the CORRECT spreadsheet (3,328 real purchases)
# ══════════════════════════════════════════════════════════════════════
print()
print("=" * 60)
print("STEP 2: Reading spreadsheet")
print(f"  File: {SPREADSHEET}")
print("=" * 60)

wb = openpyxl.load_workbook(SPREADSHEET, data_only=True)
ws = wb["Vendas"]

# Column mapping (0-indexed from iter_rows, 1-indexed for .cell()):
# Col 5  (idx 4) = cpf/cnpj
# Col 6  (idx 5) = Nome
# Col 7  (idx 6) = Cliente (email)
# Col 8  (idx 7) = Celular
# Col 9  (idx 8) = Endereco
# Col 11 (idx 10) = data_venda
# Col 12 (idx 11) = ultima_parcela
# Col 14 (idx 13) = produto
# Col 16 (idx 15) = renovacao
# Col 17 (idx 16) = nivel
# Col 22 (idx 21) = valor_total
# Col 27 (idx 26) = cancelamento

all_sales: list[dict] = []
sales_by_cpf: dict[str, list[dict]] = {}
sales_by_email: dict[str, list[dict]] = {}
sales_by_name: dict[str, list[dict]] = {}

for row in ws.iter_rows(min_row=2):
    name_val = row[5].value  # Nome
    if not name_val or not str(name_val).strip():
        continue

    name = safe_str(name_val)
    cpf = safe_str(row[4].value)  # cpf/cnpj
    email = safe_str(row[6].value)  # Cliente (email)
    celular = safe_phone(row[7].value)  # Celular
    endereco = safe_str(row[8].value)  # Endereco
    data_venda = safe_date(row[10].value)  # data_venda
    ultima_parcela = safe_date(row[11].value)  # ultima_parcela
    produto = safe_str(row[13].value)  # produto
    renovacao_str = safe_str(row[15].value)  # renovacao
    nivel = safe_str(row[16].value)  # nivel
    valor_total = safe_float(row[21].value)  # valor_total
    cancel_raw = row[26].value  # cancelamento

    # Parse cancellation
    cancelamento = False
    if cancel_raw is not None:
        cv = str(cancel_raw).strip().lower()
        cancelamento = cv in ("true", "sim", "1", "cancelado")

    sale = {
        "nome": name,
        "cpf": cpf,
        "email": email,
        "celular": celular,
        "endereco": endereco,
        "data": data_venda,
        "ultima_parcela": ultima_parcela,
        "produto": produto,
        "renovacao": renovacao_str,
        "nivel": nivel,
        "valor": valor_total,
        "cancelamento": cancelamento,
    }
    all_sales.append(sale)

    # Index by CPF (primary key for real purchase counting)
    if cpf:
        cpf_clean = cpf.replace(".", "").replace("-", "").replace("/", "").strip()
        if cpf_clean:
            sales_by_cpf.setdefault(cpf_clean, []).append(sale)

    # Index by email
    if email:
        ekey = email.lower().strip()
        if ekey:
            sales_by_email.setdefault(ekey, []).append(sale)

    # Index by normalized name
    nkey = norm(name)
    if nkey:
        sales_by_name.setdefault(nkey, []).append(sale)

wb.close()

print(f"  Loaded {len(all_sales)} sales from spreadsheet")
print(f"  Unique CPFs: {len(sales_by_cpf)}")
print(f"  Unique emails: {len(sales_by_email)}")
print(f"  Unique names: {len(sales_by_name)}")

# Count real renewals: CPFs with 2+ entries
real_renewal_cpfs = {cpf: sales for cpf, sales in sales_by_cpf.items() if len(sales) >= 2}
print(f"  Real renewals (CPFs with 2+ purchases): {len(real_renewal_cpfs)}")


# ══════════════════════════════════════════════════════════════════════
# STEP 3: Fetch Vindi customer details via API
# ══════════════════════════════════════════════════════════════════════
print()
print("=" * 60)
print("STEP 3: Fetching Vindi customer details via API")
print("=" * 60)

auth_token = base64.b64encode(f"{VINDI_KEY}:".encode()).decode()
vindi_customers: list[dict] = []

page = 1
while True:
    url = f"https://app.vindi.com.br/api/v1/customers?per_page=50&page={page}"
    req = urllib.request.Request(url)
    req.add_header("Authorization", f"Basic {auth_token}")
    req.add_header("Accept", "application/json")

    try:
        resp = urllib.request.urlopen(req, timeout=30)
        data = json.loads(resp.read())
        batch = data.get("customers", [])
        if not batch:
            break
        vindi_customers.extend(batch)
        print(f"  Page {page}: {len(batch)} customers (total so far: {len(vindi_customers)})")
        page += 1
        time.sleep(0.5)
    except Exception as e:
        print(f"  Error on page {page}: {e}")
        break

print(f"  Total Vindi customers from API: {len(vindi_customers)}")

# Build lookup by email and by name for Vindi API customers
vindi_api_by_email: dict[str, dict] = {}
vindi_api_by_name: dict[str, dict] = {}
vindi_api_by_id: dict[int, dict] = {}

for vc in vindi_customers:
    vindi_api_by_id[vc["id"]] = vc
    if vc.get("email"):
        vindi_api_by_email[vc["email"].lower().strip()] = vc
    if vc.get("name"):
        vindi_api_by_name[norm(vc["name"])] = vc


# ══════════════════════════════════════════════════════════════════════
# STEP 4: Cross-reference everything
# ══════════════════════════════════════════════════════════════════════
print()
print("=" * 60)
print("STEP 4: Cross-referencing Vindi bills <-> Spreadsheet <-> Vindi API")
print("=" * 60)

customer_data: dict[str, dict] = {}
match_planilha = 0
match_vindi_api = 0

for cliente, vdata in vindi_by_customer.items():
    # Start with Vindi bill data
    entry = {
        "nome": cliente,
        "email": "",
        "cpf": "",
        "celular": "",
        # Vindi billing data (recurrence/installments)
        "vindi_bills": len(vdata["bills"]),
        "vindi_total": round(vdata["total"], 2),
        "vindi_first": vdata["first"],
        "vindi_last": vdata["last"],
        # Spreadsheet data (real purchases) — defaults
        "compras": 0,
        "compras_detail": [],
        "produto": "",
        "nivel": "",
        "renovacao": False,
        "cancelamento": False,
        "fonte": "Vindi",
        "_bills": vdata["bills"],
    }

    # ── Try matching to spreadsheet ──
    matched_sales: list[dict] = []

    # 1. Try by email (if cliente looks like an email)
    if "@" in cliente:
        matched_sales = sales_by_email.get(cliente.lower().strip(), [])

    # 2. Try by normalized name
    if not matched_sales:
        nkey = norm(cliente)
        matched_sales = sales_by_name.get(nkey, [])

    # 3. If matched, also check if email matches for better cross-ref
    if not matched_sales:
        # Try Vindi API email -> spreadsheet
        vapi = vindi_api_by_name.get(norm(cliente))
        if vapi and vapi.get("email"):
            api_email = vapi["email"].lower().strip()
            matched_sales = sales_by_email.get(api_email, [])

    if matched_sales:
        match_planilha += 1
        entry["fonte"] = "Planilha + Vindi"

        # Count REAL purchases by CPF (the true metric)
        cpf_from_sales = matched_sales[0].get("cpf", "")
        cpf_clean = cpf_from_sales.replace(".", "").replace("-", "").replace("/", "").strip()

        # Get ALL sales for this CPF (may include sales under different names)
        if cpf_clean and cpf_clean in sales_by_cpf:
            all_cpf_sales = sales_by_cpf[cpf_clean]
        else:
            all_cpf_sales = matched_sales

        entry["compras"] = len(all_cpf_sales)
        entry["compras_detail"] = []
        for s in sorted(all_cpf_sales, key=lambda x: x["data"] or ""):
            entry["compras_detail"].append({
                "data": s["data"],
                "produto": s["produto"],
                "valor": s["valor"],
                "renovacao": s["renovacao"],
                "nivel": s.get("nivel", ""),
            })

        # Use most recent sale for main fields
        latest = sorted(all_cpf_sales, key=lambda x: x["data"] or "")[-1]
        entry["produto"] = latest["produto"]
        entry["nivel"] = latest.get("nivel", "")
        entry["cpf"] = cpf_from_sales
        entry["email"] = matched_sales[0].get("email", "")
        entry["celular"] = matched_sales[0].get("celular", "")
        entry["cancelamento"] = any(s["cancelamento"] for s in all_cpf_sales)

        # Renewal = 2+ real purchases in spreadsheet OR any sale marked "Renovação"
        entry["renovacao"] = (
            entry["compras"] >= 2
            or any(s["renovacao"] in ("Renovação", "Renovacao") for s in all_cpf_sales)
        )

    # ── Enrich from Vindi API (especially for Vindi-only customers) ──
    vapi = vindi_api_by_name.get(norm(cliente))
    if not vapi and entry["email"]:
        vapi = vindi_api_by_email.get(entry["email"].lower().strip())

    if vapi:
        if not entry["email"] and vapi.get("email"):
            entry["email"] = vapi["email"]
        if not entry["cpf"] and vapi.get("registry_code"):
            entry["cpf"] = vapi["registry_code"]
        if not entry["nome"] or entry["nome"] == cliente:
            entry["nome"] = vapi.get("name", cliente)
        # Extract phone from API
        if not entry["celular"] and vapi.get("phones"):
            phones = vapi["phones"]
            if phones:
                p = phones[0]
                entry["celular"] = f"{p.get('area_code', '')}{p.get('number', '')}"
        if not matched_sales:
            match_vindi_api += 1

    customer_data[cliente] = entry

print(f"  Matched to spreadsheet: {match_planilha}")
print(f"  Enriched from Vindi API (no spreadsheet match): {match_vindi_api}")
print(f"  Total customer entries: {len(customer_data)}")


# ══════════════════════════════════════════════════════════════════════
# STEP 5: Compute renewal stats (CORRECT numbers)
# ══════════════════════════════════════════════════════════════════════
print()
print("=" * 60)
print("STEP 5: Computing renewal stats")
print("=" * 60)

total_vindi_customers = len(customer_data)
total_planilha_sales = len(all_sales)
unique_cpfs = len(sales_by_cpf)
real_renewals = len(real_renewal_cpfs)
recurrence_customers = sum(1 for c in customer_data.values() if c["vindi_bills"] >= 2)

# Total value of real renewals (CPFs with 2+ purchases — sum all their purchases)
total_renewal_value = 0.0
for cpf, sales in real_renewal_cpfs.items():
    total_renewal_value += sum(s["valor"] for s in sales)

renewal_stats = {
    "total_customers": total_vindi_customers,
    "total_planilha": total_planilha_sales,
    "unique_cpfs": unique_cpfs,
    "real_renewals": real_renewals,
    "recurrence_customers": recurrence_customers,
    "total_renewal_value": round(total_renewal_value, 2),
}

print(f"  Stats: {json.dumps(renewal_stats, indent=2, ensure_ascii=False)}")


# ══════════════════════════════════════════════════════════════════════
# STEP 6: Write output JS
# ══════════════════════════════════════════════════════════════════════
print()
print("=" * 60)
print("STEP 6: Writing receitas_customers.js")
print("=" * 60)

# Separate bills list from customer data output
customer_output: dict[str, dict] = {}
bills_output: dict[str, list] = {}

for k, v in customer_data.items():
    bills_list = v.pop("_bills")
    customer_output[k] = v
    bills_output[k] = bills_list

out = "var customerData = " + json.dumps(customer_output, ensure_ascii=False, indent=1) + ";\n"
out += "var renewalStats = " + json.dumps(renewal_stats, ensure_ascii=False, indent=1) + ";\n"
out += "var customerBills = " + json.dumps(bills_output, ensure_ascii=False, indent=1) + ";\n"

with open(OUTPUT_JS, "w") as f:
    f.write(out)

size = os.path.getsize(OUTPUT_JS)
print(f"  Wrote {OUTPUT_JS} ({size:,} bytes)")
print()
print("Done!")
