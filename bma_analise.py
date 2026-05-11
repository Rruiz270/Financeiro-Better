#!/usr/bin/env python3
"""
BMA Solucoes/Consultoria - Complete Financial Analysis
Fetches contas a pagar, contas a receber, clients, and projects from BMA's OMIE account.
Cross-references with Better's OMIE client list.
Produces BMA_Analise_Completa.xlsx with 6 tabs.
"""

import json
import urllib.request
import time
import os
from collections import defaultdict
from datetime import datetime, date
from typing import Any

# ---------------------------------------------------------------------------
# API credentials
# ---------------------------------------------------------------------------
BMA_KEY = "4602985397010"
BMA_SECRET = "e7e7e8ebffe8c76051459f4dbbb468e5"

BETTER_KEY = "4340156993172"
BETTER_SECRET = "dd4651357eabc69d5381e8b47a293eb0"

MESES = {
    "01": "Jan", "02": "Fev", "03": "Mar", "04": "Abr",
    "05": "Mai", "06": "Jun", "07": "Jul", "08": "Ago",
    "09": "Set", "10": "Out", "11": "Nov", "12": "Dez",
}
MES_NUM = {v: int(k) for k, v in MESES.items()}

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
# OMIE helpers
# ---------------------------------------------------------------------------

def omie_call(endpoint: str, method: str, params: dict,
              app_key: str = BMA_KEY, app_secret: str = BMA_SECRET) -> dict:
    body = json.dumps({
        "call": method,
        "app_key": app_key,
        "app_secret": app_secret,
        "param": [params],
    }).encode()
    req = urllib.request.Request(
        f"https://app.omie.com.br/api/v1/{endpoint}",
        data=body,
        headers={"Content-Type": "application/json"},
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        return json.loads(resp.read())


def omie_paginate(endpoint: str, method: str, list_key: str,
                  per_page: int = 50, delay: float = 3,
                  app_key: str = BMA_KEY, app_secret: str = BMA_SECRET,
                  extra_params: dict | None = None) -> list[dict]:
    results: list[dict] = []
    page = 1
    total_pages = None
    while True:
        params = {"pagina": page, "registros_por_pagina": per_page, "apenas_importado_api": "N"}
        if extra_params:
            params.update(extra_params)
        for attempt in range(4):
            try:
                d = omie_call(endpoint, method, params, app_key, app_secret)
                break
            except Exception as e:
                wait = delay * (attempt + 1)
                print(f"  Retry {attempt+1} page {page}: {e}  (waiting {wait}s)")
                time.sleep(wait)
        else:
            print(f"  SKIP page {page} after 4 attempts")
            page += 1
            if total_pages and page > total_pages:
                break
            continue

        if total_pages is None:
            total_pages = d.get("total_de_paginas", 0)
            total_records = d.get("total_de_registros", 0)
            print(f"  {method}: {total_records} records, {total_pages} pages")

        results.extend(d.get(list_key, []))
        if page % 5 == 0:
            print(f"    Page {page}/{total_pages} ({len(results)} fetched)")
        if page >= total_pages:
            break
        page += 1
        time.sleep(delay)
    return results


def parse_br_date(dt: str) -> date | None:
    """Parse dd/mm/yyyy to date object."""
    if not dt or "/" not in dt:
        return None
    parts = dt.split("/")
    if len(parts) < 3:
        return None
    try:
        return date(int(parts[2]), int(parts[1]), int(parts[0]))
    except (ValueError, IndexError):
        return None


def month_key(d: date | None) -> str:
    """Return 'Mmm/YYYY' from a date."""
    if d is None:
        return ""
    return f"{MESES.get(f'{d.month:02d}', '???')}/{d.year}"


def sort_month_key(mk: str):
    """Sort key for 'Mmm/YYYY' strings."""
    if "/" not in mk:
        return (0, 0)
    parts = mk.split("/")
    return (int(parts[1]), MES_NUM.get(parts[0], 0))


# ---------------------------------------------------------------------------
# Excel formatting helpers (openpyxl)
# ---------------------------------------------------------------------------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
ATRASADO_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
POSITIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NEGATIVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
BRL_FMT = '#,##0.00'
DATE_FMT = 'DD/MM/YYYY'


def style_header(ws, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = ws.dimensions


def auto_width(ws, min_w: int = 10, max_w: int = 45):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


def write_rows(ws, headers: list[str], rows: list[list], money_cols: set[int] | None = None,
               date_cols: set[int] | None = None):
    """Write header + data rows with formatting."""
    money_cols = money_cols or set()
    date_cols = date_cols or set()
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = THIN_BORDER
            if ci in money_cols:
                cell.number_format = BRL_FMT
                cell.alignment = Alignment(horizontal="right")
            elif ci in date_cols:
                cell.number_format = DATE_FMT
                cell.alignment = Alignment(horizontal="center")
    style_header(ws, len(headers))
    auto_width(ws)


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    print(f"=== BMA Analise Completa - {datetime.now().strftime('%Y-%m-%d %H:%M')} ===\n")

    # -----------------------------------------------------------------------
    # 1. Fetch BMA clients
    # -----------------------------------------------------------------------
    print("[1/6] BMA - Clientes...")
    bma_clients = omie_paginate("geral/clientes/", "ListarClientes", "clientes_cadastro",
                                delay=3)
    print(f"  -> {len(bma_clients)} BMA clients\n")

    bma_cli_map: dict[int, dict] = {}
    for c in bma_clients:
        cod = c.get("codigo_cliente_omie")
        bma_cli_map[cod] = {
            "nome": c.get("nome_fantasia", "") or c.get("razao_social", ""),
            "razao": c.get("razao_social", ""),
            "cpf_cnpj": c.get("cnpj_cpf", ""),
        }

    # -----------------------------------------------------------------------
    # 2. Fetch BMA contas a pagar
    # -----------------------------------------------------------------------
    print("[2/6] BMA - Contas a Pagar...")
    time.sleep(3)
    bma_cp = omie_paginate("financas/contapagar/", "ListarContasPagar", "conta_pagar_cadastro",
                           delay=3)
    print(f"  -> {len(bma_cp)} contas a pagar\n")

    # -----------------------------------------------------------------------
    # 3. Fetch BMA contas a receber
    # -----------------------------------------------------------------------
    print("[3/6] BMA - Contas a Receber...")
    time.sleep(3)
    bma_cr = omie_paginate("financas/contareceber/", "ListarContasReceber", "conta_receber_cadastro",
                           delay=3)
    print(f"  -> {len(bma_cr)} contas a receber\n")

    # -----------------------------------------------------------------------
    # 4. Fetch BMA projects (departamentos / categorias)
    # -----------------------------------------------------------------------
    print("[4/6] BMA - Projetos...")
    time.sleep(3)
    bma_projects = []
    try:
        bma_projects = omie_paginate("geral/projetos/", "ListarProjetos", "cadastro",
                                     delay=3)
    except Exception as e:
        print(f"  Projects fetch error: {e}")
    print(f"  -> {len(bma_projects)} projects\n")

    # -----------------------------------------------------------------------
    # 5. Fetch BMA categories
    # -----------------------------------------------------------------------
    print("[5/6] BMA - Categorias...")
    time.sleep(3)
    bma_categories = []
    try:
        bma_categories = omie_paginate("geral/categorias/", "ListarCategorias", "categoria_cadastro",
                                       delay=3)
    except Exception as e:
        print(f"  Categories fetch error: {e}")
    print(f"  -> {len(bma_categories)} categories\n")

    cat_map: dict[str, str] = {}
    for cat in bma_categories:
        cod = cat.get("codigo", "")
        desc = cat.get("descricao", "")
        cat_map[cod] = desc

    # -----------------------------------------------------------------------
    # 6. Fetch Better's client list for cross-reference
    # -----------------------------------------------------------------------
    print("[6/6] Better - Clientes (cross-reference)...")
    time.sleep(3)
    better_clients = omie_paginate("geral/clientes/", "ListarClientes", "clientes_cadastro",
                                   delay=3, app_key=BETTER_KEY, app_secret=BETTER_SECRET)
    print(f"  -> {len(better_clients)} Better clients\n")

    better_names: set[str] = set()
    better_docs: set[str] = set()
    for c in better_clients:
        nf = (c.get("nome_fantasia", "") or "").strip().upper()
        rs = (c.get("razao_social", "") or "").strip().upper()
        doc = (c.get("cnpj_cpf", "") or "").strip()
        if nf:
            better_names.add(nf)
        if rs:
            better_names.add(rs)
        if doc:
            better_docs.add(doc)

    # -----------------------------------------------------------------------
    # Build project map
    # -----------------------------------------------------------------------
    proj_map: dict[int, str] = {}
    for p in bma_projects:
        pid = p.get("codigo", 0) or p.get("nCodProjeto", 0)
        pname = p.get("nome", "") or p.get("cNome", "") or p.get("descricao", "")
        if pid:
            proj_map[pid] = pname

    # -----------------------------------------------------------------------
    # Process contas a pagar
    # -----------------------------------------------------------------------
    cp_rows = []
    for c in bma_cp:
        status = c.get("status_titulo", "")
        forn_cod = c.get("codigo_cliente_fornecedor", 0)
        forn_info = bma_cli_map.get(forn_cod, {})
        forn_nome = forn_info.get("nome", str(forn_cod)) if forn_info else str(forn_cod)
        cat_code = c.get("codigo_categoria", "")
        proj_code = c.get("codigo_projeto", 0) or 0
        proj_nome = proj_map.get(proj_code, str(proj_code) if proj_code else "")
        d_emissao = parse_br_date(c.get("data_emissao", ""))
        d_vencimento = parse_br_date(c.get("data_vencimento", ""))
        d_previsao = parse_br_date(c.get("data_previsao", ""))
        valor = c.get("valor_documento", 0) or 0
        obs = c.get("observacao", "") or ""
        num_doc = c.get("numero_documento", "") or ""
        num_pedido = c.get("numero_pedido", "") or ""

        primary_date = d_previsao or d_vencimento or d_emissao
        mk = month_key(primary_date)

        cp_rows.append({
            "fornecedor": forn_nome,
            "forn_cod": forn_cod,
            "cpf_cnpj": forn_info.get("cpf_cnpj", ""),
            "data_emissao": d_emissao,
            "data_vencimento": d_vencimento,
            "data_previsao": d_previsao,
            "primary_date": primary_date,
            "valor": valor,
            "status": status,
            "categoria": cat_code,
            "cat_desc": cat_map.get(cat_code, ""),
            "projeto_cod": proj_code,
            "projeto": proj_nome,
            "obs": obs,
            "num_doc": num_doc,
            "mes": mk,
        })

    cp_rows.sort(key=lambda r: r["primary_date"] or date(1900, 1, 1), reverse=True)

    # Process contas a receber
    cr_rows = []
    for c in bma_cr:
        status = c.get("status_titulo", "")
        cli_cod = c.get("codigo_cliente_fornecedor", 0)
        cli_info = bma_cli_map.get(cli_cod, {})
        cli_nome = cli_info.get("nome", str(cli_cod)) if cli_info else str(cli_cod)
        cat_code = c.get("codigo_categoria", "")
        proj_code = c.get("codigo_projeto", 0) or 0
        proj_nome = proj_map.get(proj_code, str(proj_code) if proj_code else "")
        d_emissao = parse_br_date(c.get("data_emissao", ""))
        d_vencimento = parse_br_date(c.get("data_vencimento", ""))
        d_previsao = parse_br_date(c.get("data_previsao", ""))
        valor = c.get("valor_documento", 0) or 0

        primary_date = d_previsao or d_vencimento or d_emissao
        mk = month_key(primary_date)

        cr_rows.append({
            "cliente": cli_nome,
            "cli_cod": cli_cod,
            "data_emissao": d_emissao,
            "data_vencimento": d_vencimento,
            "data_previsao": d_previsao,
            "primary_date": primary_date,
            "valor": valor,
            "status": status,
            "categoria": cat_code,
            "cat_desc": cat_map.get(cat_code, ""),
            "projeto_cod": proj_code,
            "projeto": proj_nome,
            "mes": mk,
        })

    cr_rows.sort(key=lambda r: r["primary_date"] or date(1900, 1, 1), reverse=True)

    # -----------------------------------------------------------------------
    # Build Excel workbook
    # -----------------------------------------------------------------------
    print("Building Excel workbook...")
    wb = Workbook()

    # ===== TAB 1: Resumo Mensal =====
    ws1 = wb.active
    ws1.title = "Resumo Mensal"

    # Monthly despesas (from contas a pagar, excluding CANCELADO)
    desp_monthly: dict[str, float] = defaultdict(float)
    desp_count: dict[str, int] = defaultdict(int)
    for r in cp_rows:
        if r["status"] == "CANCELADO":
            continue
        mk = r["mes"]
        if mk:
            desp_monthly[mk] += r["valor"]
            desp_count[mk] += 1

    # Monthly receitas (from contas a receber, excluding CANCELADO)
    rec_monthly: dict[str, float] = defaultdict(float)
    rec_count: dict[str, int] = defaultdict(int)
    for r in cr_rows:
        if r["status"] == "CANCELADO":
            continue
        mk = r["mes"]
        if mk:
            rec_monthly[mk] += r["valor"]
            rec_count[mk] += 1

    all_months = sorted(
        set(list(desp_monthly.keys()) + list(rec_monthly.keys())),
        key=sort_month_key,
    )

    resumo_headers = ["Mes", "Despesas (R$)", "# Despesas", "Receitas (R$)", "# Receitas", "Resultado (R$)"]
    resumo_rows = []
    total_desp = total_rec = 0
    total_dn = total_rn = 0
    for mk in all_months:
        d = desp_monthly.get(mk, 0)
        dn = desp_count.get(mk, 0)
        r = rec_monthly.get(mk, 0)
        rn = rec_count.get(mk, 0)
        res = r - d
        resumo_rows.append([mk, d, dn, r, rn, res])
        total_desp += d
        total_rec += r
        total_dn += dn
        total_rn += rn

    resumo_rows.append(["TOTAL", total_desp, total_dn, total_rec, total_rn, total_rec - total_desp])

    write_rows(ws1, resumo_headers, resumo_rows, money_cols={2, 4, 6})

    # Color resultado column
    for ri in range(2, len(resumo_rows) + 2):
        cell = ws1.cell(row=ri, column=6)
        if cell.value and isinstance(cell.value, (int, float)):
            cell.fill = POSITIVE_FILL if cell.value >= 0 else NEGATIVE_FILL

    # Bold total row
    total_row_idx = len(resumo_rows) + 1
    for ci in range(1, 7):
        ws1.cell(row=total_row_idx, column=ci).font = Font(bold=True, size=11)

    # ===== TAB 2: Contas a Pagar =====
    ws2 = wb.create_sheet("Contas a Pagar")
    cp_headers = [
        "Fornecedor", "CPF/CNPJ", "Data Emissao", "Data Vencimento", "Data Previsao",
        "Mes", "Valor (R$)", "Status", "Categoria", "Cat. Descricao", "Projeto", "Obs", "Num Doc"
    ]
    cp_excel = []
    for r in cp_rows:
        cp_excel.append([
            r["fornecedor"], r["cpf_cnpj"],
            r["data_emissao"], r["data_vencimento"], r["data_previsao"],
            r["mes"], r["valor"], r["status"],
            r["categoria"], r["cat_desc"], r["projeto"],
            r["obs"][:100] if r["obs"] else "", r["num_doc"],
        ])
    write_rows(ws2, cp_headers, cp_excel, money_cols={7}, date_cols={3, 4, 5})

    # Highlight atrasados
    for ri in range(2, len(cp_excel) + 2):
        if ws2.cell(row=ri, column=8).value == "ATRASADO":
            for ci in range(1, len(cp_headers) + 1):
                ws2.cell(row=ri, column=ci).fill = ATRASADO_FILL

    # ===== TAB 3: Contas a Receber =====
    ws3 = wb.create_sheet("Contas a Receber")
    cr_headers = [
        "Cliente", "Data Emissao", "Data Vencimento", "Data Previsao",
        "Mes", "Valor (R$)", "Status", "Categoria", "Cat. Descricao", "Projeto"
    ]
    cr_excel = []
    for r in cr_rows:
        cr_excel.append([
            r["cliente"],
            r["data_emissao"], r["data_vencimento"], r["data_previsao"],
            r["mes"], r["valor"], r["status"],
            r["categoria"], r["cat_desc"], r["projeto"],
        ])
    write_rows(ws3, cr_headers, cr_excel, money_cols={6}, date_cols={2, 3, 4})

    for ri in range(2, len(cr_excel) + 2):
        if ws3.cell(row=ri, column=7).value == "ATRASADO":
            for ci in range(1, len(cr_headers) + 1):
                ws3.cell(row=ri, column=ci).fill = ATRASADO_FILL

    # ===== TAB 4: Atrasados =====
    ws4 = wb.create_sheet("Atrasados")
    atr_headers = [
        "Tipo", "Nome", "Data Vencimento", "Data Previsao", "Mes",
        "Valor (R$)", "Status", "Categoria", "Cat. Descricao", "Projeto"
    ]
    atr_rows_data = []
    for r in cp_rows:
        if r["status"] == "ATRASADO":
            atr_rows_data.append([
                "Pagar", r["fornecedor"], r["data_vencimento"], r["data_previsao"],
                r["mes"], r["valor"], r["status"], r["categoria"], r["cat_desc"], r["projeto"],
            ])
    atr_cp_total = sum(r[5] for r in atr_rows_data)

    for r in cr_rows:
        if r["status"] == "ATRASADO":
            atr_rows_data.append([
                "Receber", r["cliente"], r["data_vencimento"], r["data_previsao"],
                r["mes"], r["valor"], r["status"], r["categoria"], r["cat_desc"], r["projeto"],
            ])
    atr_cr_total = sum(r[5] for r in atr_rows_data if r[0] == "Receber")

    write_rows(ws4, atr_headers, atr_rows_data, money_cols={6}, date_cols={3, 4})

    # Color by type
    for ri in range(2, len(atr_rows_data) + 2):
        tipo = ws4.cell(row=ri, column=1).value
        fill = ATRASADO_FILL if tipo == "Pagar" else PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        for ci in range(1, len(atr_headers) + 1):
            ws4.cell(row=ri, column=ci).fill = fill

    # ===== TAB 5: Fornecedores =====
    ws5 = wb.create_sheet("Fornecedores")
    forn_agg: dict[str, dict] = {}
    for r in cp_rows:
        if r["status"] == "CANCELADO":
            continue
        nome = r["fornecedor"]
        if nome not in forn_agg:
            forn_agg[nome] = {
                "total": 0, "count": 0, "categories": set(),
                "cpf_cnpj": r["cpf_cnpj"], "forn_cod": r["forn_cod"],
            }
        forn_agg[nome]["total"] += r["valor"]
        forn_agg[nome]["count"] += 1
        if r["categoria"]:
            forn_agg[nome]["categories"].add(r["categoria"])

    forn_sorted = sorted(forn_agg.items(), key=lambda x: x[1]["total"], reverse=True)

    forn_headers = [
        "Fornecedor", "CPF/CNPJ", "Total (R$)", "Qtd Pagamentos",
        "Categorias", "Existe na Better?", "Tipo (PJ/CLT/Beneficio)"
    ]
    forn_rows_data = []
    for nome, info in forn_sorted:
        doc = info["cpf_cnpj"]
        # Cross-reference: check by document or name
        in_better = "Sim" if (doc and doc in better_docs) else ""
        if not in_better:
            upper_nome = nome.strip().upper()
            for bn in better_names:
                if upper_nome in bn or bn in upper_nome:
                    in_better = "Sim (nome)"
                    break

        # Classify: PJ, CLT, beneficio
        cats = info["categories"]
        tipo = ""
        cat_strs = ", ".join(sorted(cats))
        has_salario = any(c.startswith("2.03.01") for c in cats)
        has_pj = any(c.startswith("2.03.94") or c.startswith("2.16") for c in cats)
        has_beneficio = any("flash" in nome.lower() or "pluxee" in nome.lower()
                           or "sodexo" in nome.lower() or "alelo" in nome.lower()
                           or c.startswith("2.03.02") or c.startswith("2.03.03")
                           for c in cats)
        if has_salario:
            tipo = "CLT"
        elif has_pj:
            tipo = "PJ"
        elif has_beneficio:
            tipo = "Beneficio (VR/VT)"
        elif any(c.startswith("2.03.") for c in cats):
            tipo = "Pessoal"

        forn_rows_data.append([
            nome, doc, info["total"], info["count"],
            cat_strs, in_better, tipo,
        ])

    write_rows(ws5, forn_headers, forn_rows_data, money_cols={3})

    # ===== TAB 6: Projetos =====
    ws6 = wb.create_sheet("Projetos")
    proj_agg: dict[str, dict] = {}
    for r in cp_rows:
        if r["status"] == "CANCELADO":
            continue
        pname = r["projeto"] or "(Sem Projeto)"
        if pname not in proj_agg:
            proj_agg[pname] = {"total": 0, "count": 0, "fornecedores": set(), "categories": set()}
        proj_agg[pname]["total"] += r["valor"]
        proj_agg[pname]["count"] += 1
        proj_agg[pname]["fornecedores"].add(r["fornecedor"])
        if r["categoria"]:
            proj_agg[pname]["categories"].add(r["categoria"])

    proj_sorted = sorted(proj_agg.items(), key=lambda x: x[1]["total"], reverse=True)

    proj_headers = ["Projeto", "Total (R$)", "Qtd Pagamentos", "Fornecedores", "Categorias"]
    proj_rows_data = []
    for pname, info in proj_sorted:
        proj_rows_data.append([
            pname, info["total"], info["count"],
            ", ".join(sorted(info["fornecedores"]))[:200],
            ", ".join(sorted(info["categories"])),
        ])

    write_rows(ws6, proj_headers, proj_rows_data, money_cols={2})

    # -----------------------------------------------------------------------
    # Save workbook
    # -----------------------------------------------------------------------
    out_path = os.path.join(OUTPUT_DIR, "BMA_Analise_Completa.xlsx")
    wb.save(out_path)
    print(f"\nSaved: {out_path}\n")

    # -----------------------------------------------------------------------
    # Print summary analysis
    # -----------------------------------------------------------------------
    print("=" * 80)
    print("SUMMARY ANALYSIS - BMA Solucoes/Consultoria")
    print("=" * 80)

    print(f"\n--- Record Counts ---")
    print(f"  Contas a Pagar:  {len(bma_cp)}")
    print(f"  Contas a Receber: {len(bma_cr)}")
    print(f"  Clientes BMA:    {len(bma_clients)}")
    print(f"  Projetos:        {len(bma_projects)}")
    print(f"  Categorias:      {len(bma_categories)}")
    print(f"  Clientes Better: {len(better_clients)}")

    print(f"\n--- Monthly Totals (all non-cancelled) ---")
    print(f"  Total Despesas:  R$ {total_desp:,.2f}")
    print(f"  Total Receitas:  R$ {total_rec:,.2f}")
    print(f"  Resultado:       R$ {total_rec - total_desp:,.2f}")

    # Atrasados
    print(f"\n--- ATRASADOS ---")
    cp_atrasado = [r for r in cp_rows if r["status"] == "ATRASADO"]
    cr_atrasado = [r for r in cr_rows if r["status"] == "ATRASADO"]
    print(f"  Contas a Pagar ATRASADO:  {len(cp_atrasado)} records, R$ {sum(r['valor'] for r in cp_atrasado):,.2f}")
    print(f"  Contas a Receber ATRASADO: {len(cr_atrasado)} records, R$ {sum(r['valor'] for r in cr_atrasado):,.2f}")
    if cr_atrasado:
        print(f"\n  Detalhamento Contas a Receber Atrasadas:")
        for r in sorted(cr_atrasado, key=lambda x: x["valor"], reverse=True):
            print(f"    {r['cliente']:<40s}  R$ {r['valor']:>12,.2f}  venc: {r['data_vencimento']}  cat: {r['categoria']}")

    # PJ vs CLT classification
    print(f"\n--- PJ vs CLT Analysis ---")
    pj_people: list[tuple[str, float, int]] = []
    clt_people: list[tuple[str, float, int]] = []
    benefit_entries: list[tuple[str, float, int]] = []

    for nome, info in forn_sorted:
        cats = info["categories"]
        has_salario = any(c.startswith("2.03.01") for c in cats)
        has_pj = any(c.startswith("2.03.94") or c.startswith("2.16") for c in cats)
        is_benefit = ("flash" in nome.lower() or "pluxee" in nome.lower()
                      or "sodexo" in nome.lower() or "alelo" in nome.lower())

        if is_benefit:
            benefit_entries.append((nome, info["total"], info["count"]))
        elif has_salario:
            clt_people.append((nome, info["total"], info["count"]))
        elif has_pj:
            pj_people.append((nome, info["total"], info["count"]))

    print(f"\n  CLT (salarios - cat 2.03.01): {len(clt_people)} fornecedores")
    for name, total, count in sorted(clt_people, key=lambda x: x[1], reverse=True):
        print(f"    {name:<40s}  R$ {total:>12,.2f}  ({count} pgtos)")

    print(f"\n  PJ (cat 2.03.94 / 2.16.xx): {len(pj_people)} fornecedores")
    for name, total, count in sorted(pj_people, key=lambda x: x[1], reverse=True):
        print(f"    {name:<40s}  R$ {total:>12,.2f}  ({count} pgtos)")

    # Flash App
    print(f"\n--- Flash App / Benefits ---")
    for name, total, count in benefit_entries:
        print(f"    {name:<40s}  R$ {total:>12,.2f}  ({count} pgtos)")
    flash_items = [r for r in cp_rows if "flash" in r["fornecedor"].lower() and r["status"] != "CANCELADO"]
    if flash_items:
        print(f"\n  Flash by month:")
        flash_by_month: dict[str, float] = defaultdict(float)
        for r in flash_items:
            if r["mes"]:
                flash_by_month[r["mes"]] += r["valor"]
        for mk in sorted(flash_by_month.keys(), key=sort_month_key):
            print(f"    {mk}: R$ {flash_by_month[mk]:,.2f}")

    # Helen
    print(f"\n--- Helen's Payments ---")
    helen_items = [r for r in cp_rows if "helen" in r["fornecedor"].lower() and r["status"] != "CANCELADO"]
    if helen_items:
        helen_total = sum(r["valor"] for r in helen_items)
        print(f"  {len(helen_items)} payments, total R$ {helen_total:,.2f}")
        for r in sorted(helen_items, key=lambda x: x["primary_date"] or date(1900, 1, 1), reverse=True)[:20]:
            print(f"    {r['fornecedor']:<40s}  R$ {r['valor']:>10,.2f}  {r['data_previsao']}  cat: {r['categoria']}  proj: {r['projeto']}")
    else:
        print("  No payments found with 'Helen' in fornecedor name.")
        # Search broader
        helen_like = [r for r in cp_rows if any(h in r["fornecedor"].lower() for h in ["helen", "helena"]) and r["status"] != "CANCELADO"]
        if helen_like:
            print(f"  Found {len(helen_like)} with 'helena':")
            for r in helen_like[:10]:
                print(f"    {r['fornecedor']}  R$ {r['valor']:,.2f}")

    # Raphael Lima and Alice
    print(f"\n--- Raphael Lima & Alice ---")
    for search_name in ["raphael lima", "alice"]:
        items = [r for r in cp_rows if search_name in r["fornecedor"].lower() and r["status"] != "CANCELADO"]
        if items:
            total_val = sum(r["valor"] for r in items)
            print(f"\n  '{search_name}': {len(items)} payments, total R$ {total_val:,.2f}")
            for r in sorted(items, key=lambda x: x["primary_date"] or date(1900, 1, 1), reverse=True)[:15]:
                print(f"    {r['fornecedor']:<40s}  R$ {r['valor']:>10,.2f}  {r['data_previsao']}  cat: {r['categoria']}  proj: {r['projeto']}")
        else:
            print(f"\n  '{search_name}': No payments found. Searching partial...")
            partial = [r for r in cp_rows if search_name.split()[0] in r["fornecedor"].lower() and r["status"] != "CANCELADO"]
            if partial:
                for r in partial[:5]:
                    print(f"    {r['fornecedor']}  R$ {r['valor']:,.2f}  cat: {r['categoria']}")

    # Category breakdown
    print(f"\n--- Top Categories by Spend ---")
    cat_totals: dict[str, float] = defaultdict(float)
    for r in cp_rows:
        if r["status"] != "CANCELADO" and r["categoria"]:
            label = f"{r['categoria']} ({r['cat_desc']})" if r["cat_desc"] else r["categoria"]
            cat_totals[label] += r["valor"]
    for label, total in sorted(cat_totals.items(), key=lambda x: x[1], reverse=True)[:20]:
        print(f"    {label:<50s}  R$ {total:>12,.2f}")

    print(f"\n{'=' * 80}")
    print(f"Excel saved to: {out_path}")
    print(f"{'=' * 80}")


if __name__ == "__main__":
    main()
