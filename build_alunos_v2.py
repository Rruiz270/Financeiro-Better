#!/usr/bin/env python3
"""
build_alunos_v2.py
==================
Rebuilds the unified student database from:
  1. Vendas tab (3,328 sales) in the master spreadsheet
  2. Alunos tab (3,569 student records) in the master spreadsheet
  3. Vindi paid bills (7,642 records) from receitas_all_data.js

Outputs:
  - alunos_data.js          (JS var for the HTML dashboard)
  - Alunos_Base_Completa.xlsx (6-tab Excel workbook)
  - Updates alunos.html       (field-name alignment if needed)

Run:  python3 build_alunos_v2.py
"""

import json
import os
import re
import warnings
from collections import defaultdict
from datetime import datetime, date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ─── Paths ──────────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.abspath(__file__))
XLSX_SRC = "/Users/Raphael/Downloads/Vendas para Emissão de NF (5).xlsx"
VINDI_JS = os.path.join(BASE, "receitas_all_data.js")
OUT_JS = os.path.join(BASE, "alunos_data.js")
OUT_XLSX = os.path.join(BASE, "Alunos_Base_Completa.xlsx")
OUT_HTML = os.path.join(BASE, "alunos.html")

TODAY = date(2026, 5, 12)
TODAY_STR = TODAY.isoformat()


# ═══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def safe_str(v):
    """Convert value to stripped string, or empty string if None."""
    if v is None:
        return ""
    return str(v).strip()


def normalize_cpf(raw):
    """Digits-only CPF/CNPJ, handling floats and strings."""
    if raw is None:
        return ""
    s = str(raw).strip()
    # Remove trailing .0 from float representations
    if s.endswith(".0"):
        s = s[:-2]
    # Keep only digits
    s = re.sub(r"[^\d]", "", s)
    return s


def normalize_email(raw):
    """Lowercase stripped email."""
    if raw is None:
        return ""
    return str(raw).strip().lower()


def normalize_celular(raw):
    """Convert float/int phone to string, strip .0, keep digits only."""
    if raw is None:
        return ""
    s = str(raw).strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = re.sub(r"[^\d]", "", s)
    return s


def safe_date(v):
    """Extract a date string (YYYY-MM-DD) from datetime, string, or None."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    # Try ISO parse
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except (ValueError, TypeError):
            continue
    # If it looks like a date serial number (all digits, length > 6), skip
    if re.match(r"^\d{7,}$", s):
        return ""
    return ""


def safe_float(v):
    """Convert value to float, 0.0 on failure."""
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def safe_bool(v):
    """Check if value is truthy / True / 'True'."""
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in ("true", "1", "sim", "yes", "x")


def normalize_name(raw):
    """Normalize a name for matching: lowercase, strip accents approx, collapse spaces."""
    if not raw:
        return ""
    import unicodedata
    s = str(raw).strip()
    # NFKD decompose then strip combining marks
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).lower().strip()


def classify_modalidade(produto, doc_tipo):
    """Classify the student's modality based on product name and document type."""
    if not produto:
        p = ""
    else:
        p = str(produto).lower()

    if "flow" in p:
        return "Community Flow"
    if "particular" in p:
        return "Private"
    if "espanhol" in p or "spanish" in p:
        return "Espanhol"
    if "faap" in p:
        return "In-Company (FAAP)"
    if "imers" in p:  # Imersão / Imersao
        return "Imersao"
    if doc_tipo and str(doc_tipo).upper() == "CNPJ":
        return "In-Company"
    return "Community"


def determine_status(cancelamento, data_fim_str, vindi_last_str, vindi_only=False):
    """
    Determine student status.
    Priority:
      1. cancelamento flag  -> "Cancelado"
      2. data_fim >= today  -> "Ativo"
      3. vindi_last >= 2026-03-01 -> "Ativo (Vindi)"
      4. data_fim < today, no recent vindi -> "Expirado"
      5. vindi-only, last < 2025-06-01 -> "Inativo"
    """
    if cancelamento:
        return "Cancelado"

    if data_fim_str:
        try:
            fim = datetime.strptime(data_fim_str, "%Y-%m-%d").date()
            if fim >= TODAY:
                return "Ativo"
        except (ValueError, TypeError):
            pass

    if vindi_last_str:
        try:
            vl = datetime.strptime(vindi_last_str, "%Y-%m-%d").date()
            if vl >= date(2026, 3, 1):
                return "Ativo (Vindi)"
            if vindi_only and vl < date(2025, 6, 1):
                return "Inativo"
        except (ValueError, TypeError):
            pass

    if data_fim_str:
        try:
            fim = datetime.strptime(data_fim_str, "%Y-%m-%d").date()
            if fim < TODAY:
                return "Expirado"
        except (ValueError, TypeError):
            pass

    if vindi_only:
        return "Inativo"

    return "Expirado"


# ═══════════════════════════════════════════════════════════════════════════════
#  1. READ VENDAS TAB
# ═══════════════════════════════════════════════════════════════════════════════

def read_vendas():
    """Read Vendas tab. Returns list of dicts, one per sale."""
    print("[1/4] Reading Vendas tab...")
    wb = openpyxl.load_workbook(XLSX_SRC, data_only=True)
    ws = wb["Vendas"]

    sales = []
    for r in range(2, ws.max_row + 1):
        sale_key = ws.cell(r, 1).value
        if sale_key is None:
            continue  # skip empty rows at bottom

        cpf_raw = ws.cell(r, 5).value
        cpf = normalize_cpf(cpf_raw)
        nome = safe_str(ws.cell(r, 6).value)
        email = normalize_email(ws.cell(r, 7).value)
        celular = normalize_celular(ws.cell(r, 8).value)
        endereco = safe_str(ws.cell(r, 9).value)
        data_transacao = safe_date(ws.cell(r, 10).value)
        data_venda = safe_date(ws.cell(r, 11).value)
        ultima_parcela = safe_date(ws.cell(r, 12).value)
        forma = safe_str(ws.cell(r, 13).value)
        produto = safe_str(ws.cell(r, 14).value)
        fonte = safe_str(ws.cell(r, 15).value)
        renovacao = safe_str(ws.cell(r, 16).value)
        nivel = safe_str(ws.cell(r, 17).value)
        duracao = safe_float(ws.cell(r, 19).value)
        valor_total = safe_float(ws.cell(r, 22).value)
        cancelamento = safe_bool(ws.cell(r, 27).value)
        data_cancelamento = safe_date(ws.cell(r, 32).value)
        doc_tipo = safe_str(ws.cell(r, 4).value)  # CPF or CNPJ

        sales.append({
            "sale_key": sale_key,
            "cpf": cpf,
            "nome": nome,
            "email": email,
            "celular": celular,
            "endereco": endereco,
            "data_transacao": data_transacao,
            "data_venda": data_venda,
            "ultima_parcela": ultima_parcela,
            "forma": forma,
            "produto": produto,
            "fonte_vendedor": fonte,
            "renovacao": renovacao,
            "nivel": nivel,
            "duracao": duracao,
            "valor_total": valor_total,
            "cancelamento": cancelamento,
            "data_cancelamento": data_cancelamento,
            "doc_tipo": doc_tipo.upper() if doc_tipo else "CPF",
        })

    wb.close()
    print(f"   -> {len(sales)} sales read")
    return sales


# ═══════════════════════════════════════════════════════════════════════════════
#  2. READ ALUNOS TAB
# ═══════════════════════════════════════════════════════════════════════════════

def read_alunos_tab():
    """
    Read Alunos tab. Returns list of dicts.
    Columns: #, NOME, EMAIL, CELULAR, STATUS, TIPO, EMPRESA, PRODUTO,
             Horas ofertadas, INICIO, FIM, (None), LEVELING, NIVEL, ...
    """
    print("[2/4] Reading Alunos tab...")
    wb = openpyxl.load_workbook(XLSX_SRC, data_only=True)
    ws = wb["Alunos"]

    alunos = []
    for r in range(2, ws.max_row + 1):
        nome = safe_str(ws.cell(r, 2).value)
        if not nome:
            continue

        email = normalize_email(ws.cell(r, 3).value)
        celular = normalize_celular(ws.cell(r, 4).value)
        status = safe_str(ws.cell(r, 5).value).upper()
        tipo = safe_str(ws.cell(r, 6).value)  # B2C, B2B, BOLSISTA, B2B2C
        empresa = safe_str(ws.cell(r, 7).value)
        produto = safe_str(ws.cell(r, 8).value)
        inicio = safe_date(ws.cell(r, 10).value)
        fim = safe_date(ws.cell(r, 11).value)
        nivel = safe_str(ws.cell(r, 14).value)

        alunos.append({
            "nome": nome,
            "email": email,
            "celular": celular,
            "status_alunos": status,
            "tipo": tipo,
            "empresa": empresa,
            "produto": produto,
            "inicio": inicio,
            "fim": fim,
            "nivel": nivel,
        })

    wb.close()
    print(f"   -> {len(alunos)} student records read")
    return alunos


# ═══════════════════════════════════════════════════════════════════════════════
#  3. READ VINDI BILLS
# ═══════════════════════════════════════════════════════════════════════════════

def read_vindi():
    """
    Read receitas_all_data.js. Returns dict keyed by normalized cliente name
    with: bills (count), total, first_date, last_date, details list.
    """
    print("[3/4] Reading Vindi bills...")
    with open(VINDI_JS, "r", encoding="utf-8") as f:
        content = f.read()

    # Strip "var data = " prefix and trailing ";"
    json_str = content.replace("var data = ", "", 1).rstrip().rstrip(";")
    records = json.loads(json_str)
    print(f"   -> {len(records)} Vindi bill records loaded")

    # Group by cliente name (normalized)
    grouped = defaultdict(lambda: {
        "bills": 0,
        "total": 0.0,
        "first": None,
        "last": None,
        "details": [],
        "cliente_original": "",
    })

    for rec in records:
        cliente = safe_str(rec.get("cliente", ""))
        if not cliente:
            continue
        key = normalize_name(cliente)
        g = grouped[key]
        g["cliente_original"] = cliente
        g["bills"] += 1
        g["total"] += safe_float(rec.get("valor", 0))

        d = safe_str(rec.get("data", ""))
        if d:
            if g["first"] is None or d < g["first"]:
                g["first"] = d
            if g["last"] is None or d > g["last"]:
                g["last"] = d

        g["details"].append({
            "data": d,
            "valor": safe_float(rec.get("valor", 0)),
            "situacao": safe_str(rec.get("situacao", "")),
            "mes": safe_str(rec.get("mes", "")),
            "categoria": safe_str(rec.get("categoria", "")),
        })

    print(f"   -> {len(grouped)} unique Vindi customers")
    return grouped


# ═══════════════════════════════════════════════════════════════════════════════
#  4. CONSOLIDATE
# ═══════════════════════════════════════════════════════════════════════════════

def consolidate(sales, alunos_tab, vindi):
    """
    Merge all three sources into a unified student list.
    Primary key: CPF (digits). Fallback: email (lowercase).
    """
    print("[4/4] Consolidating records...")

    # ── Group sales by CPF (or email if no CPF) ────────────────────────────
    cpf_groups = defaultdict(list)
    email_groups = defaultdict(list)

    for s in sales:
        key = s["cpf"] if s["cpf"] else s["email"]
        if not key:
            key = f"__nokey_{s['sale_key']}"
        if s["cpf"]:
            cpf_groups[s["cpf"]].append(s)
        elif s["email"]:
            email_groups[s["email"]].append(s)
        else:
            cpf_groups[key].append(s)

    # ── Index Alunos tab by email for enrichment ──────────────────────────
    alunos_by_email = {}
    for a in alunos_tab:
        if a["email"]:
            alunos_by_email[a["email"]] = a

    # ── Index Vindi by normalized name ────────────────────────────────────
    vindi_matched = set()

    def find_vindi(nome, email):
        """Try to match a student to Vindi data by name or email."""
        # Try by normalized name first
        nname = normalize_name(nome)
        if nname and nname in vindi:
            vindi_matched.add(nname)
            return vindi[nname]
        # Try email prefix as name (some Vindi entries use email)
        if email:
            ekey = normalize_name(email.split("@")[0].replace(".", " "))
            if ekey and ekey in vindi:
                vindi_matched.add(ekey)
                return vindi[ekey]
        return None

    # ── Build unified records ─────────────────────────────────────────────
    students = {}  # keyed by cpf or email

    # Process CPF-grouped sales
    for cpf, group in cpf_groups.items():
        group.sort(key=lambda s: s["data_venda"] or "")
        latest = group[-1]
        earliest = group[0]

        # Pick best data from latest sale, fallback from earlier
        nome = latest["nome"] or earliest["nome"]
        email = latest["email"] or earliest["email"]
        celular = latest["celular"] or earliest["celular"]
        endereco = latest["endereco"] or earliest["endereco"]

        # Calculate total from spreadsheet
        total_planilha = sum(s["valor_total"] for s in group)

        # Check cancelamento: if ALL sales cancelled
        any_cancel = any(s["cancelamento"] for s in group)
        all_cancel = all(s["cancelamento"] for s in group)
        latest_cancel = latest["cancelamento"]

        # Dates
        data_inicio = earliest["data_venda"]
        data_fim = latest["ultima_parcela"] or latest["data_venda"]

        # Best nivel
        nivel = latest["nivel"]
        for s in reversed(group):
            if s["nivel"]:
                nivel = s["nivel"]
                break

        # Renovacao count
        renov_flags = [s for s in group if s["renovacao"] and s["renovacao"].lower() not in ("", "-", "nova", "new")]
        compras = len(group)

        # Doc tipo
        doc_tipo = latest["doc_tipo"]
        modalidade = classify_modalidade(latest["produto"], doc_tipo)

        # Enrich from Alunos tab
        aluno_info = alunos_by_email.get(email, {})
        if aluno_info:
            if not nivel and aluno_info.get("nivel"):
                nivel = aluno_info["nivel"]
            if not celular and aluno_info.get("celular"):
                celular = aluno_info["celular"]
            # Use Alunos tab fim if newer
            a_fim = aluno_info.get("fim", "")
            if a_fim and (not data_fim or a_fim > data_fim):
                data_fim = a_fim

        # Vindi match
        vindi_info = find_vindi(nome, email)
        vindi_bills = 0
        vindi_total = 0.0
        vindi_first = ""
        vindi_last = ""
        vindi_details = []
        if vindi_info:
            vindi_bills = vindi_info["bills"]
            vindi_total = round(vindi_info["total"], 2)
            vindi_first = vindi_info["first"] or ""
            vindi_last = vindi_info["last"] or ""
            vindi_details = vindi_info["details"]

        # Status
        status = determine_status(
            cancelamento=latest_cancel if not (compras > 1 and not all_cancel) else all_cancel,
            data_fim_str=data_fim,
            vindi_last_str=vindi_last,
        )

        # Build vendas detail for modal
        vendas_detail = []
        for s in group:
            vendas_detail.append({
                "produto": s["produto"],
                "data_venda": s["data_venda"],
                "valor": s["valor_total"],
                "renovacao": s["renovacao"],
                "vendedor": s["fonte_vendedor"],
                "forma": s["forma"],
                "duracao": s["duracao"],
                "cancelamento": s["cancelamento"],
                "data_cancelamento": s["data_cancelamento"],
            })

        rec = {
            "nome": nome,
            "email": email,
            "cpf": cpf if not cpf.startswith("__nokey_") else "",
            "celular": celular,
            "endereco": endereco,
            "tipo_doc": doc_tipo,
            "tipo_cliente": "PJ" if doc_tipo == "CNPJ" else "PF",
            "modalidade": modalidade,
            "nivel": nivel,
            "produto_principal": latest["produto"],
            "data_inicio": data_inicio,
            "data_fim": data_fim,
            "meses_curso": latest["duracao"],
            "valor_planilha": round(total_planilha, 2),
            "total_gasto": round(total_planilha, 2),
            "renovacao": latest["renovacao"],
            "vendedor": latest["fonte_vendedor"],
            "cancelamento": latest_cancel,
            "cancelou": latest_cancel,
            "data_cancelamento": latest["data_cancelamento"] if latest_cancel else "",
            "status": status,
            "compras": compras,
            "renovacoes": len(renov_flags),
            "contrato_ativo": status in ("Ativo", "Ativo (Vindi)"),
            "aulas_remanescentes": 0,
            "vendas": vendas_detail,
            "vindi_bills": vindi_bills,
            "vindi_total": vindi_total,
            "vindi_first": vindi_first,
            "vindi_last": vindi_last,
            "vindi_last_payment": vindi_last,
            "vindi_bill_detail": vindi_details,
            "fonte_planilha": True,
            "fonte_vindi": vindi_bills > 0,
            "fonte": "Ambos" if vindi_bills > 0 else "Planilha",
        }

        students[cpf] = rec

    # Process email-grouped sales (no CPF)
    for email, group in email_groups.items():
        if email in students:
            # Merge into existing
            existing = students[email]
            for s in group:
                existing["vendas"].append({
                    "produto": s["produto"],
                    "data_venda": s["data_venda"],
                    "valor": s["valor_total"],
                    "renovacao": s["renovacao"],
                    "vendedor": s["fonte_vendedor"],
                    "forma": s["forma"],
                    "duracao": s["duracao"],
                    "cancelamento": s["cancelamento"],
                    "data_cancelamento": s["data_cancelamento"],
                })
                existing["compras"] += 1
                existing["valor_planilha"] += s["valor_total"]
                existing["total_gasto"] += s["valor_total"]
            continue

        group.sort(key=lambda s: s["data_venda"] or "")
        latest = group[-1]
        earliest = group[0]
        nome = latest["nome"] or earliest["nome"]
        celular = latest["celular"] or earliest["celular"]
        endereco = latest["endereco"] or earliest["endereco"]
        total_planilha = sum(s["valor_total"] for s in group)
        compras = len(group)
        nivel = latest["nivel"]
        for s in reversed(group):
            if s["nivel"]:
                nivel = s["nivel"]
                break
        renov_flags = [s for s in group if s["renovacao"] and s["renovacao"].lower() not in ("", "-", "nova", "new")]
        doc_tipo = latest["doc_tipo"]
        modalidade = classify_modalidade(latest["produto"], doc_tipo)

        # Alunos tab
        aluno_info = alunos_by_email.get(email, {})
        if aluno_info:
            if not nivel and aluno_info.get("nivel"):
                nivel = aluno_info["nivel"]
            if not celular and aluno_info.get("celular"):
                celular = aluno_info["celular"]

        data_inicio = earliest["data_venda"]
        data_fim = latest["ultima_parcela"] or latest["data_venda"]
        if aluno_info:
            a_fim = aluno_info.get("fim", "")
            if a_fim and (not data_fim or a_fim > data_fim):
                data_fim = a_fim

        vindi_info = find_vindi(nome, email)
        vindi_bills = vindi_info["bills"] if vindi_info else 0
        vindi_total = round(vindi_info["total"], 2) if vindi_info else 0.0
        vindi_first = (vindi_info["first"] or "") if vindi_info else ""
        vindi_last = (vindi_info["last"] or "") if vindi_info else ""
        vindi_details = vindi_info["details"] if vindi_info else []

        latest_cancel = latest["cancelamento"]
        all_cancel = all(s["cancelamento"] for s in group)
        status = determine_status(
            cancelamento=latest_cancel if compras == 1 else all_cancel,
            data_fim_str=data_fim,
            vindi_last_str=vindi_last,
        )

        vendas_detail = []
        for s in group:
            vendas_detail.append({
                "produto": s["produto"],
                "data_venda": s["data_venda"],
                "valor": s["valor_total"],
                "renovacao": s["renovacao"],
                "vendedor": s["fonte_vendedor"],
                "forma": s["forma"],
                "duracao": s["duracao"],
                "cancelamento": s["cancelamento"],
                "data_cancelamento": s["data_cancelamento"],
            })

        rec = {
            "nome": nome,
            "email": email,
            "cpf": "",
            "celular": celular,
            "endereco": endereco,
            "tipo_doc": doc_tipo,
            "tipo_cliente": "PJ" if doc_tipo == "CNPJ" else "PF",
            "modalidade": modalidade,
            "nivel": nivel,
            "produto_principal": latest["produto"],
            "data_inicio": data_inicio,
            "data_fim": data_fim,
            "meses_curso": latest["duracao"],
            "valor_planilha": round(total_planilha, 2),
            "total_gasto": round(total_planilha, 2),
            "renovacao": latest["renovacao"],
            "vendedor": latest["fonte_vendedor"],
            "cancelamento": latest_cancel,
            "cancelou": latest_cancel,
            "data_cancelamento": latest["data_cancelamento"] if latest_cancel else "",
            "status": status,
            "compras": compras,
            "renovacoes": len(renov_flags),
            "contrato_ativo": status in ("Ativo", "Ativo (Vindi)"),
            "aulas_remanescentes": 0,
            "vendas": vendas_detail,
            "vindi_bills": vindi_bills,
            "vindi_total": vindi_total,
            "vindi_first": vindi_first,
            "vindi_last": vindi_last,
            "vindi_last_payment": vindi_last,
            "vindi_bill_detail": vindi_details,
            "fonte_planilha": True,
            "fonte_vindi": vindi_bills > 0,
            "fonte": "Ambos" if vindi_bills > 0 else "Planilha",
        }

        students[email] = rec

    # ── Vindi-only records (not matched to any sale) ──────────────────────
    vindi_only_count = 0
    for nname, vinfo in vindi.items():
        if nname in vindi_matched:
            continue

        cliente = vinfo["cliente_original"]
        vindi_only_count += 1

        status = determine_status(
            cancelamento=False,
            data_fim_str="",
            vindi_last_str=vinfo["last"] or "",
            vindi_only=True,
        )

        key = f"__vindi_{nname}"
        rec = {
            "nome": cliente,
            "email": "",
            "cpf": "",
            "celular": "",
            "endereco": "",
            "tipo_doc": "",
            "tipo_cliente": "PF",
            "modalidade": "Community",
            "nivel": "",
            "produto_principal": "",
            "data_inicio": vinfo["first"] or "",
            "data_fim": vinfo["last"] or "",
            "meses_curso": 0,
            "valor_planilha": 0.0,
            "total_gasto": 0.0,
            "renovacao": "",
            "vendedor": "",
            "cancelamento": False,
            "cancelou": False,
            "data_cancelamento": "",
            "status": status,
            "compras": 0,
            "renovacoes": 0,
            "contrato_ativo": status in ("Ativo", "Ativo (Vindi)"),
            "aulas_remanescentes": 0,
            "vendas": [],
            "vindi_bills": vinfo["bills"],
            "vindi_total": round(vinfo["total"], 2),
            "vindi_first": vinfo["first"] or "",
            "vindi_last": vinfo["last"] or "",
            "vindi_last_payment": vinfo["last"] or "",
            "vindi_bill_detail": vinfo["details"],
            "fonte_planilha": False,
            "fonte_vindi": True,
            "fonte": "Vindi",
        }
        students[key] = rec

    all_students = list(students.values())
    all_students.sort(key=lambda s: (s["nome"] or "").lower())

    # ── Stats ─────────────────────────────────────────────────────────────
    planilha_recs = [s for s in all_students if s["fonte_planilha"]]
    vindi_recs = [s for s in all_students if s["fonte_vindi"]]
    matched = [s for s in all_students if s["fonte_planilha"] and s["fonte_vindi"]]

    stats = {
        "total": len(all_students),
        "ativos": len([s for s in all_students if "Ativo" in s["status"]]),
        "cancelados": len([s for s in all_students if s["status"] == "Cancelado"]),
        "expirados": len([s for s in all_students if s["status"] == "Expirado"]),
        "inativos": len([s for s in all_students if s["status"] == "Inativo"]),
        "community": len([s for s in all_students if s["modalidade"] == "Community"]),
        "flow": len([s for s in all_students if s["modalidade"] == "Community Flow"]),
        "espanhol": len([s for s in all_students if s["modalidade"] == "Espanhol"]),
        "private": len([s for s in all_students if s["modalidade"] == "Private"]),
        "in_company": len([s for s in all_students if "In-Company" in s["modalidade"]]),
        "imersao": len([s for s in all_students if s["modalidade"] == "Imersao"]),
        "renovacoes": len([s for s in all_students if s["compras"] > 1]),
        "pf": len([s for s in all_students if s["tipo_cliente"] == "PF"]),
        "pj": len([s for s in all_students if s["tipo_cliente"] == "PJ"]),
        "total_planilha": len(planilha_recs),
        "total_vindi": len(vindi_recs),
        "planilha_only": len([s for s in all_students if s["fonte_planilha"] and not s["fonte_vindi"]]),
        "vindi_only": vindi_only_count,
        "matched": len(matched),
    }

    print(f"   -> {len(all_students)} consolidated students")
    print(f"      Planilha: {stats['total_planilha']} | Vindi: {stats['total_vindi']} | Matched: {stats['matched']} | Vindi-only: {stats['vindi_only']}")
    print(f"      Ativos: {stats['ativos']} | Cancelados: {stats['cancelados']} | Expirados: {stats['expirados']} | Inativos: {stats['inativos']}")

    return all_students, stats


# ═══════════════════════════════════════════════════════════════════════════════
#  OUTPUT: alunos_data.js
# ═══════════════════════════════════════════════════════════════════════════════

def write_js(students, stats):
    """Write alunos_data.js with all student records and stats."""
    print("Writing alunos_data.js...")

    # Serialize — we omit vindi_bill_detail from JS to keep file size manageable
    # Actually the HTML expects it, so we include it but truncate to last 24 entries
    js_records = []
    for s in students:
        rec = dict(s)
        # Truncate Vindi detail to last 24 for JS performance
        if len(rec.get("vindi_bill_detail", [])) > 24:
            rec["vindi_bill_detail"] = rec["vindi_bill_detail"][-24:]
        js_records.append(rec)

    js_str = json.dumps(js_records, ensure_ascii=False, indent=1)
    stats_str = json.dumps(stats, ensure_ascii=False, indent=1)

    with open(OUT_JS, "w", encoding="utf-8") as f:
        f.write(f"var alunosData = {js_str};\n")
        f.write(f"var alunosStats = {stats_str};\n")

    size_mb = os.path.getsize(OUT_JS) / (1024 * 1024)
    print(f"   -> alunos_data.js written ({size_mb:.1f} MB, {len(students)} records)")


# ═══════════════════════════════════════════════════════════════════════════════
#  OUTPUT: Alunos_Base_Completa.xlsx
# ═══════════════════════════════════════════════════════════════════════════════

# Styles
HEADER_FILL = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

STATUS_FILLS = {
    "Ativo": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
    "Ativo (Vindi)": PatternFill(start_color="B2DFDB", end_color="B2DFDB", fill_type="solid"),
    "Cancelado": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
    "Expirado": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
    "Inativo": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
}


def style_header(ws, col_count):
    """Apply header styling to row 1."""
    for c in range(1, col_count + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, col_count, max_width=40):
    """Auto-fit column widths based on content."""
    for c in range(1, col_count + 1):
        max_len = 0
        for r in range(1, min(ws.max_row + 1, 200)):
            val = ws.cell(r, c).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        width = min(max_len + 3, max_width)
        ws.column_dimensions[get_column_letter(c)].width = max(width, 10)


MAIN_HEADERS = [
    "Nome", "Email", "CPF/CNPJ", "Celular", "Endereco", "Tipo Doc",
    "Tipo Cliente", "Modalidade", "Nivel", "Produto", "Data Inicio",
    "Data Fim", "Meses", "Valor Planilha (R$)", "Renovacao", "Vendedor",
    "Cancelamento", "Data Cancel.", "Status", "Compras", "Renovacoes",
    "Vindi Bills", "Vindi Total (R$)", "Vindi Primeiro", "Vindi Ultimo", "Fonte"
]


def write_student_row(ws, row, s):
    """Write one student record to a worksheet row."""
    vals = [
        s["nome"], s["email"], s["cpf"], s["celular"], s["endereco"],
        s["tipo_doc"], s["tipo_cliente"], s["modalidade"], s["nivel"],
        s["produto_principal"], s["data_inicio"], s["data_fim"],
        s["meses_curso"], s["valor_planilha"], s["renovacao"], s["vendedor"],
        "Sim" if s["cancelamento"] else "Nao",
        s["data_cancelamento"], s["status"], s["compras"], s["renovacoes"],
        s["vindi_bills"], s["vindi_total"], s["vindi_first"], s["vindi_last"],
        s["fonte"],
    ]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row, c, v)
        cell.border = THIN_BORDER
        if c in (14, 23):  # money columns
            cell.number_format = '#,##0.00'
        if c == 19:  # status column
            fill = STATUS_FILLS.get(str(v), None)
            if fill:
                cell.fill = fill


def build_sheet(wb, name, students):
    """Create a sheet with full student data."""
    ws = wb.create_sheet(name)
    # Headers
    for c, h in enumerate(MAIN_HEADERS, 1):
        ws.cell(1, c, h)
    style_header(ws, len(MAIN_HEADERS))

    for i, s in enumerate(students, 2):
        write_student_row(ws, i, s)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(MAIN_HEADERS))}{len(students)+1}"
    auto_width(ws, len(MAIN_HEADERS))
    return ws


def write_xlsx(students, stats):
    """Write Alunos_Base_Completa.xlsx with 6 tabs."""
    print("Writing Alunos_Base_Completa.xlsx...")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # 1. Base Completa
    all_sorted = sorted(students, key=lambda s: (s["nome"] or "").lower())
    build_sheet(wb, "Base Completa", all_sorted)

    # 2. Ativos
    ativos = [s for s in all_sorted if "Ativo" in s["status"]]
    build_sheet(wb, "Ativos", ativos)

    # 3. Cancelados
    cancelados = [s for s in all_sorted if s["status"] == "Cancelado"]
    build_sheet(wb, "Cancelados", cancelados)

    # 4. Renovacoes (CPFs with 2+ purchases)
    renovacoes = [s for s in all_sorted if s["compras"] > 1]
    build_sheet(wb, "Renovacoes", renovacoes)

    # 5. Para Importacao Portal (active students, subset of columns)
    ws_portal = wb.create_sheet("Para Importacao Portal")
    portal_headers = ["Nome", "Email", "CPF/CNPJ", "Celular", "Modalidade", "Nivel", "Status"]
    for c, h in enumerate(portal_headers, 1):
        ws_portal.cell(1, c, h)
    style_header(ws_portal, len(portal_headers))

    portal_students = [s for s in all_sorted if "Ativo" in s["status"]]
    for i, s in enumerate(portal_students, 2):
        vals = [s["nome"], s["email"], s["cpf"], s["celular"], s["modalidade"], s["nivel"], s["status"]]
        for c, v in enumerate(vals, 1):
            cell = ws_portal.cell(i, c, v)
            cell.border = THIN_BORDER
    ws_portal.auto_filter.ref = f"A1:{get_column_letter(len(portal_headers))}{len(portal_students)+1}"
    auto_width(ws_portal, len(portal_headers))

    # 6. Resumo (stats)
    ws_resumo = wb.create_sheet("Resumo")
    resumo_data = [
        ("Metrica", "Valor"),
        ("Total Alunos (consolidado)", stats["total"]),
        ("Ativos", stats["ativos"]),
        ("Cancelados", stats["cancelados"]),
        ("Expirados", stats["expirados"]),
        ("Inativos", stats["inativos"]),
        ("", ""),
        ("MODALIDADES", ""),
        ("Community", stats["community"]),
        ("Community Flow", stats["flow"]),
        ("Espanhol", stats["espanhol"]),
        ("Private (Particular)", stats["private"]),
        ("In-Company", stats["in_company"]),
        ("Imersao", stats["imersao"]),
        ("", ""),
        ("RENOVACOES", ""),
        ("Alunos com 2+ compras", stats["renovacoes"]),
        ("", ""),
        ("TIPO DOCUMENTO", ""),
        ("Pessoa Fisica (PF)", stats["pf"]),
        ("Pessoa Juridica (PJ)", stats["pj"]),
        ("", ""),
        ("FONTES DE DADOS", ""),
        ("Total na Planilha", stats["total_planilha"]),
        ("Total no Vindi", stats["total_vindi"]),
        ("Apenas Planilha", stats["planilha_only"]),
        ("Apenas Vindi", stats["vindi_only"]),
        ("Cruzados (ambos)", stats["matched"]),
        ("", ""),
        ("Gerado em", TODAY_STR),
    ]

    for r, (label, val) in enumerate(resumo_data, 1):
        ws_resumo.cell(r, 1, label)
        ws_resumo.cell(r, 2, val)
        if r == 1:
            ws_resumo.cell(r, 1).font = Font(bold=True, size=12)
            ws_resumo.cell(r, 2).font = Font(bold=True, size=12)
        ws_resumo.cell(r, 1).border = THIN_BORDER
        ws_resumo.cell(r, 2).border = THIN_BORDER

    ws_resumo.column_dimensions["A"].width = 30
    ws_resumo.column_dimensions["B"].width = 15

    # Header row styling for Resumo
    ws_resumo.cell(1, 1).fill = HEADER_FILL
    ws_resumo.cell(1, 1).font = HEADER_FONT
    ws_resumo.cell(1, 2).fill = HEADER_FILL
    ws_resumo.cell(1, 2).font = HEADER_FONT

    wb.save(OUT_XLSX)
    size_mb = os.path.getsize(OUT_XLSX) / (1024 * 1024)
    print(f"   -> Alunos_Base_Completa.xlsx written ({size_mb:.1f} MB)")
    print(f"      Tabs: Base Completa ({len(all_sorted)}), Ativos ({len(ativos)}), "
          f"Cancelados ({len(cancelados)}), Renovacoes ({len(renovacoes)}), "
          f"Portal ({len(portal_students)}), Resumo")


# ═══════════════════════════════════════════════════════════════════════════════
#  OUTPUT: Update alunos.html
# ═══════════════════════════════════════════════════════════════════════════════

def update_html(stats):
    """
    Update alunos.html to:
    - Add 'Imersao' to modalidade filter
    - Add Inativo tab
    - Ensure status badges handle 'Ativo (Vindi)' and 'Inativo'
    - Update badge CSS for new statuses
    """
    print("Updating alunos.html...")

    with open(OUT_HTML, "r", encoding="utf-8") as f:
        html = f.read()

    # 1. Add 'Imersao' to modalidade filter if missing
    if 'value="Imersao"' not in html:
        html = html.replace(
            '<option value="In-Company">In-Company</option>\n</select>',
            '<option value="In-Company">In-Company</option>\n'
            '<option value="In-Company (FAAP)">In-Company (FAAP)</option>\n'
            '<option value="Imersao">Imersao</option>\n</select>'
        )

    # 2. Add 'Inativo' tab if missing
    if 'data-tab="inativos"' not in html:
        html = html.replace(
            '<button class="tab" data-tab="vindi-only"',
            '<button class="tab" data-tab="inativos" id="tab-inativos">Inativos <span class="count" id="tc-inativos">0</span></button>\n'
            '<button class="tab" data-tab="vindi-only"'
        )

    # 3. Add badge-inativo and badge-vindi CSS if missing
    if ".badge-inativo" not in html:
        html = html.replace(
            ".badge-concluido{background:#172554;color:#60a5fa}",
            ".badge-concluido{background:#172554;color:#60a5fa}\n"
            ".badge-inativo{background:#1e293b;color:#94a3b8}\n"
            ".badge-vindi{background:#064e3b;color:#34d399}"
        )

    # 4. Update renderSummary to show Inativos count from stats
    # Replace the existing s-total-all card with inativos
    if 's-inativos' not in html:
        html = html.replace(
            '<div class="card blue"><div class="val" id="s-total-all">-</div><div class="lbl">Total (incl. Vindi)</div></div>',
            '<div class="card blue"><div class="val" id="s-inativos">-</div><div class="lbl">Inativos</div></div>'
        )

    # 5. Update JS: renderSummary function to use stats object and new field
    # Replace the renderSummary function
    old_render_summary = '''function renderSummary() {
 var planilha = data.filter(function(a){return a.fonte_planilha});
 var all = data;
 var ativos = planilha.filter(function(a){return a.status==="Ativo"}).length;
 var cancelados = planilha.filter(function(a){return a.status==="Cancelado"}).length;
 var expirados = planilha.filter(function(a){return a.status==="Expirado"}).length;
 var renovacoes = 0;
 planilha.forEach(function(a){renovacoes += a.renovacoes});

 document.getElementById("s-total").textContent = planilha.length.toLocaleString("pt-BR");
 document.getElementById("s-ativos").textContent = ativos.toLocaleString("pt-BR");
 document.getElementById("s-cancelados").textContent = cancelados.toLocaleString("pt-BR");
 document.getElementById("s-expirados").textContent = expirados.toLocaleString("pt-BR");
 document.getElementById("s-renovacoes").textContent = renovacoes.toLocaleString("pt-BR");
 document.getElementById("s-total-all").textContent = all.length.toLocaleString("pt-BR");
 document.getElementById("total-label").textContent = planilha.length.toLocaleString("pt-BR");
}'''

    new_render_summary = '''function renderSummary() {
 var all = data;
 var ativos = all.filter(function(a){return a.status&&a.status.indexOf("Ativo")!==-1}).length;
 var cancelados = all.filter(function(a){return a.status==="Cancelado"}).length;
 var expirados = all.filter(function(a){return a.status==="Expirado"}).length;
 var inativos = all.filter(function(a){return a.status==="Inativo"}).length;
 var renovacoes = all.filter(function(a){return a.compras>1}).length;

 document.getElementById("s-total").textContent = all.length.toLocaleString("pt-BR");
 document.getElementById("s-ativos").textContent = ativos.toLocaleString("pt-BR");
 document.getElementById("s-cancelados").textContent = cancelados.toLocaleString("pt-BR");
 document.getElementById("s-expirados").textContent = expirados.toLocaleString("pt-BR");
 document.getElementById("s-renovacoes").textContent = renovacoes.toLocaleString("pt-BR");
 document.getElementById("s-inativos").textContent = inativos.toLocaleString("pt-BR");
 document.getElementById("total-label").textContent = all.length.toLocaleString("pt-BR");
}'''

    html = html.replace(old_render_summary, new_render_summary)

    # 6. Update renderModalidade to count all students, not just planilha
    html = html.replace(
        'var planilha = data.filter(function(a){return a.fonte_planilha});\n var mods = {};\n planilha.forEach',
        'var allRecs = data;\n var mods = {};\n allRecs.forEach'
    )

    # 7. Update updateTabCounts for new tabs and correct status matching
    old_tab_counts = '''function updateTabCounts() {
 var planilha = data.filter(function(a){return a.fonte_planilha});
 var vindiOnly = data.filter(function(a){return !a.fonte_planilha});
 document.getElementById("tc-todos").textContent = planilha.length;
 document.getElementById("tc-ativos").textContent = planilha.filter(function(a){return a.status==="Ativo"}).length;
 document.getElementById("tc-cancelados").textContent = planilha.filter(function(a){return a.status==="Cancelado"}).length;
 document.getElementById("tc-expirados").textContent = planilha.filter(function(a){return a.status==="Expirado"}).length;
 var renov = planilha.filter(function(a){return a.renovacoes > 0});
 document.getElementById("tc-renovacoes").textContent = renov.length;
 document.getElementById("tc-vindi").textContent = vindiOnly.length;
}'''

    new_tab_counts = '''function updateTabCounts() {
 var all = data;
 var vindiOnly = data.filter(function(a){return !a.fonte_planilha});
 document.getElementById("tc-todos").textContent = all.length;
 document.getElementById("tc-ativos").textContent = all.filter(function(a){return a.status&&a.status.indexOf("Ativo")!==-1}).length;
 document.getElementById("tc-cancelados").textContent = all.filter(function(a){return a.status==="Cancelado"}).length;
 document.getElementById("tc-expirados").textContent = all.filter(function(a){return a.status==="Expirado"}).length;
 document.getElementById("tc-renovacoes").textContent = all.filter(function(a){return a.compras>1}).length;
 if(document.getElementById("tc-inativos")) document.getElementById("tc-inativos").textContent = all.filter(function(a){return a.status==="Inativo"}).length;
 document.getElementById("tc-vindi").textContent = vindiOnly.length;
}'''

    html = html.replace(old_tab_counts, new_tab_counts)

    # 8. Update applyFilters to handle all tabs correctly (use all data, not just planilha)
    old_filter_base = ''' var base;
 if(currentTab === "todos") base = data.filter(function(a){return a.fonte_planilha});
 else if(currentTab === "ativos") base = data.filter(function(a){return a.fonte_planilha && a.status==="Ativo"});
 else if(currentTab === "cancelados") base = data.filter(function(a){return a.fonte_planilha && a.status==="Cancelado"});
 else if(currentTab === "expirados") base = data.filter(function(a){return a.fonte_planilha && a.status==="Expirado"});
 else if(currentTab === "renovacoes") base = data.filter(function(a){return a.fonte_planilha && a.renovacoes > 0});
 else if(currentTab === "vindi-only") base = data.filter(function(a){return !a.fonte_planilha});
 else base = data.filter(function(a){return a.fonte_planilha});'''

    new_filter_base = ''' var base;
 if(currentTab === "todos") base = data.slice();
 else if(currentTab === "ativos") base = data.filter(function(a){return a.status&&a.status.indexOf("Ativo")!==-1});
 else if(currentTab === "cancelados") base = data.filter(function(a){return a.status==="Cancelado"});
 else if(currentTab === "expirados") base = data.filter(function(a){return a.status==="Expirado"});
 else if(currentTab === "renovacoes") base = data.filter(function(a){return a.compras>1});
 else if(currentTab === "inativos") base = data.filter(function(a){return a.status==="Inativo"});
 else if(currentTab === "vindi-only") base = data.filter(function(a){return !a.fonte_planilha});
 else base = data.slice();'''

    html = html.replace(old_filter_base, new_filter_base)

    # 9. Update status badge logic in renderTable to handle "Ativo (Vindi)" and "Inativo"
    old_status_logic = '''  var statusClass = "";
  if(a.status === "Ativo") statusClass = "badge-ativo";
  else if(a.status === "Cancelado") statusClass = "badge-cancelado";
  else if(a.status === "Expirado") statusClass = "badge-expirado";
  else statusClass = "badge-concluido";'''

    new_status_logic = '''  var statusClass = "";
  if(a.status && a.status.indexOf("Ativo")!==-1) statusClass = "badge-ativo";
  else if(a.status === "Cancelado") statusClass = "badge-cancelado";
  else if(a.status === "Expirado") statusClass = "badge-expirado";
  else if(a.status === "Inativo") statusClass = "badge-inativo";
  else statusClass = "badge-concluido";'''

    # Apply to both renderTable and openModal occurrences
    html = html.replace(old_status_logic, new_status_logic)
    # The modal also has a similar block - check if there's a second one
    # (It uses the same exact code pattern, so replace_all handled it above)

    # 10. Update filter count logic
    old_filter_count = ''' var totalBase = currentTab === "vindi-only" ? data.filter(function(a){return !a.fonte_planilha}).length : data.filter(function(a){return a.fonte_planilha}).length;'''
    new_filter_count = ''' var totalBase = data.length;'''
    html = html.replace(old_filter_count, new_filter_count)

    # 11. Update the modalidade cards to include Imersao in order array
    html = html.replace(
        'var order = ["Community","Community Flow","Espanhol","Private","In-Company","Desconhecido"];',
        'var order = ["Community","Community Flow","Espanhol","Private","In-Company","In-Company (FAAP)","Imersao","Desconhecido"];'
    )

    # 12. Update "renovacoes" references in table to "compras"
    html = html.replace(
        "'+(a.renovacoes||0)+'",
        "'+(a.compras||0)+'"
    )

    # 13. Update the header "Renov." column to "Compras"
    html = html.replace(
        '<th data-col="renovacoes">Renov.</th>',
        '<th data-col="compras">Compras</th>'
    )

    # 14. Update the table sort default
    html = html.replace(
        'var sortCol = "nome";',
        'var sortCol = "nome";'
    )

    # 15. Increase table row limit for better viewing
    html = html.replace(
        'var limit = Math.min(filtered.length, 500);',
        'var limit = Math.min(filtered.length, 1000);'
    )

    with open(OUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)

    print("   -> alunos.html updated")


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  build_alunos_v2.py — Rebuild Student Database")
    print(f"  Date: {TODAY_STR}")
    print("=" * 60)
    print()

    # Read sources
    sales = read_vendas()
    alunos_tab = read_alunos_tab()
    vindi = read_vindi()

    # Consolidate
    students, stats = consolidate(sales, alunos_tab, vindi)

    # Write outputs
    write_js(students, stats)
    write_xlsx(students, stats)
    update_html(stats)

    print()
    print("=" * 60)
    print("  DONE!")
    print(f"  Total students: {stats['total']}")
    print(f"  Ativos: {stats['ativos']} | Cancelados: {stats['cancelados']}")
    print(f"  Expirados: {stats['expirados']} | Inativos: {stats['inativos']}")
    print("=" * 60)


if __name__ == "__main__":
    main()
