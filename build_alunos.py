#!/usr/bin/env python3
"""
Build comprehensive student database from THREE data sources:
1. Alunos tab (217 records) — has inicio/fim dates, nivel, meses, sale_key
2. Vendas tab (229 records) — has CPF, endereco, produto, valor, renovacao, cancelamento
3. Vindi receitas_all_data.js (7,642 paid bills, 2,852 unique clients)

Links Alunos→Vendas via sale_key, then matches Vindi by email + fuzzy name.

Outputs:
- alunos_data.js  (for HTML dashboard)
- Alunos_Base_Completa.xlsx (6-tab Excel)
"""

import json
import re
import os
import unicodedata
from datetime import datetime, date
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TODAY = date(2026, 5, 12)
PROJECT = "/Users/Raphael/Financeiro-Better"
SALES_FILE = "/Users/Raphael/Downloads/Vendas para Emissão de NF - ATUALIZADA_03.xlsx"

# ─── Helpers ──────────────────────────────────────────────────────────────────

def normalize_cpf(raw):
    """Strip dots, dashes, slashes, spaces. Return digits only or empty string."""
    if raw is None:
        return ""
    s = re.sub(r"[.\-/ ]", "", str(raw).strip())
    return s if s.isdigit() else ""


def safe_date(val):
    """Convert to date object from datetime, string, or return None."""
    if val is None or val == "" or val == "-" or val == "nan":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, (int, float)):
        return None
    if isinstance(val, str):
        val = val.strip()
        if val in ("-", "", "nan", "None", "NaT"):
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


def safe_float(val):
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def safe_int(val):
    if val is None:
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def title_name(name):
    """Title-case a name but respect common exceptions."""
    if not name:
        return ""
    parts = name.strip().split()
    result = []
    lower_words = {"de", "da", "do", "das", "dos", "e", "em", "del", "di"}
    for i, p in enumerate(parts):
        if i > 0 and p.lower() in lower_words:
            result.append(p.lower())
        else:
            result.append(p.capitalize())
    return " ".join(result)


def strip_accents(s):
    """Remove accents for fuzzy matching."""
    if not s:
        return ""
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def normalize_name_for_match(name):
    """Lowercase, strip accents, remove punctuation for fuzzy matching."""
    if not name:
        return ""
    s = strip_accents(name).lower().strip()
    s = re.sub(r"[^a-z0-9 ]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def classify_modalidade(produto, tipo_doc):
    """Classify modality based on product name and document type."""
    if not produto:
        if tipo_doc == "CNPJ":
            return "In-Company"
        return "Community"
    p = produto.lower()
    if "particular" in p or "private" in p:
        return "Private"
    if "flow" in p:
        return "Community Flow"
    if "espanhol" in p or "spanish" in p:
        return "Espanhol"
    if tipo_doc == "CNPJ":
        return "In-Company"
    return "Community"


def determine_status(cancelou, data_fim, vindi_last, fonte_planilha):
    """
    Status logic:
    - Cancelado: has cancelamento in planilha
    - Ativo: data_fim >= today and not cancelled
    - Ativo (Vindi): expirado in planilha but has Vindi payment after 2026-03-01
    - Expirado: data_fim < today, no recent Vindi payment
    - Inativo: Vindi-only with last payment before 2025-06-01
    """
    vl = safe_date(vindi_last) if isinstance(vindi_last, str) else vindi_last
    vindi_recent = vl is not None and vl >= date(2026, 3, 1)
    vindi_old = vl is not None and vl < date(2025, 6, 1)

    if cancelou and fonte_planilha:
        return "Cancelado"

    df = safe_date(data_fim) if isinstance(data_fim, str) else data_fim

    if fonte_planilha:
        if df and df >= TODAY and not cancelou:
            return "Ativo"
        if vindi_recent:
            return "Ativo (Vindi)"
        return "Expirado"

    # Vindi-only
    if vindi_recent:
        return "Ativo (Vindi)"
    if vindi_old:
        return "Inativo"
    return "Expirado"


def estimate_remaining_classes(data_inicio, duracao_meses, status):
    """Estimate remaining classes. Community ~8/month, Private ~4/month."""
    if "Ativo" not in status or not data_inicio or not duracao_meses:
        return 0
    di = safe_date(data_inicio) if isinstance(data_inicio, str) else data_inicio
    if not di:
        return 0
    dur = int(duracao_meses) if duracao_meses else 12
    total_classes = dur * 8
    elapsed_months = (TODAY.year - di.year) * 12 + (TODAY.month - di.month)
    used = max(0, elapsed_months) * 8
    remaining = max(0, total_classes - used)
    return remaining


def compute_meses_restantes(data_fim, status):
    """Months remaining from today to end date."""
    if "Ativo" not in status:
        return 0
    df = safe_date(data_fim) if isinstance(data_fim, str) else data_fim
    if not df or df < TODAY:
        return 0
    return max(0, (df.year - TODAY.year) * 12 + (df.month - TODAY.month))


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 1: Read Alunos tab (217 records)
# ═══════════════════════════════════════════════════════════════════════════════

print("=" * 70)
print("STEP 1: Reading Alunos tab...")
wb = openpyxl.load_workbook(SALES_FILE, data_only=True)
ws_alunos = wb["Alunos"]

alunos_by_salekey = {}   # sale_key -> aluno record
alunos_by_email = {}     # email -> aluno record

alunos_count = 0
for r in range(2, ws_alunos.max_row + 1):
    nome = safe_str(ws_alunos.cell(r, 2).value)
    if not nome:
        continue

    aluno_id = safe_int(ws_alunos.cell(r, 1).value)
    email = safe_str(ws_alunos.cell(r, 3).value).strip().lower()
    tipo = safe_str(ws_alunos.cell(r, 4).value)  # B2C / B2B
    sale_key = safe_int(ws_alunos.cell(r, 5).value)
    celular = safe_str(ws_alunos.cell(r, 6).value)
    meses = safe_int(ws_alunos.cell(r, 7).value)
    inicio = safe_date(ws_alunos.cell(r, 8).value)
    fim = safe_date(ws_alunos.cell(r, 9).value)
    nivel = safe_str(ws_alunos.cell(r, 10).value)

    rec = {
        "aluno_id": aluno_id,
        "nome": title_name(nome),
        "email": email,
        "tipo": tipo,
        "sale_key": sale_key,
        "celular": celular,
        "meses": meses,
        "inicio": inicio,
        "fim": fim,
        "nivel": nivel,
    }

    if sale_key:
        alunos_by_salekey[sale_key] = rec
    if email:
        alunos_by_email[email] = rec

    alunos_count += 1

print(f"  Alunos loaded: {alunos_count}")
print(f"  Unique sale_keys: {len(alunos_by_salekey)}")
print(f"  Unique emails: {len(alunos_by_email)}")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 2: Read Vendas tab (229 records)
# ═══════════════════════════════════════════════════════════════════════════════

print("\nSTEP 2: Reading Vendas tab...")
ws_vendas = wb["Vendas"]

vendas_by_salekey = {}  # sale_key -> venda record
vendas_records = []

for r in range(2, ws_vendas.max_row + 1):
    sale_key_raw = ws_vendas.cell(r, 1).value
    if sale_key_raw is None:
        continue

    sale_key = safe_int(sale_key_raw)
    nf_produto = safe_str(ws_vendas.cell(r, 2).value)
    nf_servico = safe_str(ws_vendas.cell(r, 3).value)
    documento_tipo = safe_str(ws_vendas.cell(r, 4).value)  # CPF or CNPJ
    cpf_cnpj = safe_str(ws_vendas.cell(r, 5).value)
    cpf = normalize_cpf(cpf_cnpj)
    nome = safe_str(ws_vendas.cell(r, 6).value)
    email = safe_str(ws_vendas.cell(r, 7).value).strip().lower()
    celular = safe_str(ws_vendas.cell(r, 8).value)
    endereco = safe_str(ws_vendas.cell(r, 9).value)
    data_transacao = safe_date(ws_vendas.cell(r, 10).value)
    data_venda = safe_date(ws_vendas.cell(r, 11).value)
    ultima_parcela = safe_date(ws_vendas.cell(r, 12).value)
    forma = safe_str(ws_vendas.cell(r, 13).value)
    produto = safe_str(ws_vendas.cell(r, 14).value)
    fonte_vendedor = safe_str(ws_vendas.cell(r, 15).value)
    renovacao = safe_str(ws_vendas.cell(r, 16).value)
    nivel = safe_str(ws_vendas.cell(r, 17).value)
    desconto = safe_float(ws_vendas.cell(r, 18).value)
    duracao_curso = safe_str(ws_vendas.cell(r, 19).value)
    valor_total = safe_float(ws_vendas.cell(r, 22).value)
    cancel_raw = ws_vendas.cell(r, 27).value
    cancelamento = bool(cancel_raw) and str(cancel_raw).strip() not in ("-", "", "False", "0", "Não", "None")
    data_cancelamento = safe_date(ws_vendas.cell(r, 32).value)

    # Determine doc type
    if "cnpj" in documento_tipo.lower() or (cpf and len(cpf) > 11):
        tipo_doc = "CNPJ"
    elif "cpf" in documento_tipo.lower() or (cpf and len(cpf) == 11):
        tipo_doc = "CPF"
    else:
        tipo_doc = "CPF"

    rec = {
        "sale_key": sale_key,
        "cpf": cpf,
        "tipo_doc": tipo_doc,
        "nome": title_name(nome),
        "email": email,
        "celular": celular,
        "endereco": endereco,
        "data_transacao": data_transacao,
        "data_venda": data_venda,
        "ultima_parcela": ultima_parcela,
        "forma": forma,
        "produto": produto,
        "vendedor": fonte_vendedor,
        "renovacao": renovacao,
        "nivel": nivel,
        "desconto": desconto,
        "duracao_curso": duracao_curso,
        "valor_total": valor_total,
        "cancelamento": cancelamento,
        "data_cancelamento": data_cancelamento,
    }

    vendas_by_salekey[sale_key] = rec
    vendas_records.append(rec)

print(f"  Vendas loaded: {len(vendas_records)}")
print(f"  Unique sale_keys in Vendas: {len(vendas_by_salekey)}")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 3: Link Alunos to Vendas via sale_key
# ═══════════════════════════════════════════════════════════════════════════════

print("\nSTEP 3: Linking Alunos → Vendas via sale_key...")

# Build unified planilha students keyed by email (primary key for dedup)
planilha_students = {}  # email -> student dict

linked = 0
unlinked = 0

for sale_key, aluno in alunos_by_salekey.items():
    venda = vendas_by_salekey.get(sale_key)
    email = aluno["email"]

    if not email:
        # Use nome-based key as fallback
        email = aluno["nome"].lower().replace(" ", "_") + "@noemail"

    if venda:
        linked += 1
        cpf = venda["cpf"]
        tipo_doc = venda["tipo_doc"]
        endereco = venda["endereco"]
        produto = venda["produto"]
        vendedor = venda["vendedor"]
        renovacao_str = venda["renovacao"]
        valor_total = venda["valor_total"]
        forma = venda["forma"]
        cancelamento = venda["cancelamento"]
        data_cancelamento = venda["data_cancelamento"]
        data_venda = venda["data_venda"]
        duracao_curso = venda["duracao_curso"]
        celular = venda["celular"] or aluno["celular"]
        nivel = venda["nivel"] or aluno["nivel"]
    else:
        unlinked += 1
        cpf = ""
        tipo_doc = "CPF"
        endereco = ""
        produto = ""
        vendedor = ""
        renovacao_str = ""
        valor_total = 0.0
        forma = ""
        cancelamento = False
        data_cancelamento = None
        data_venda = aluno["inicio"]
        duracao_curso = str(aluno["meses"]) if aluno["meses"] else ""
        celular = aluno["celular"]
        nivel = aluno["nivel"]

    # Build venda detail for the modal
    venda_detail = None
    if venda:
        venda_detail = {
            "data_venda": str(data_venda) if data_venda else "",
            "produto": produto,
            "nivel": nivel,
            "valor": valor_total,
            "renovacao": renovacao_str,
            "vendedor": vendedor,
            "cancelamento": cancelamento,
            "data_cancel": str(data_cancelamento) if data_cancelamento else "",
            "forma": forma,
            "desconto": venda["desconto"],
            "duracao": duracao_curso,
        }

    if email in planilha_students:
        # Merge: add another sale
        student = planilha_students[email]
        if venda_detail:
            student["vendas"].append(venda_detail)
        student["total_gasto"] += valor_total
        if renovacao_str in ("Renovação", "Sim"):
            student["renovacoes"] += 1
        if cancelamento:
            student["cancelou"] = True
            if data_cancelamento:
                student["data_cancelamento"] = str(data_cancelamento)
        # Update date range from Alunos tab (authoritative)
        if aluno["inicio"]:
            existing_inicio = safe_date(student["data_inicio"])
            if not existing_inicio or aluno["inicio"] < existing_inicio:
                student["data_inicio"] = str(aluno["inicio"])
        if aluno["fim"]:
            existing_fim = safe_date(student["data_fim"])
            if not existing_fim or aluno["fim"] > existing_fim:
                student["data_fim"] = str(aluno["fim"])
        # Prefer non-empty
        if not student["cpf"] and cpf:
            student["cpf"] = cpf
        if not student["celular"] and celular:
            student["celular"] = celular
        if not student["endereco"] and endereco:
            student["endereco"] = endereco
        if nivel:
            student["nivel"] = nivel
        if produto:
            student["produto_principal"] = produto
        if not student["tipo_doc"] or student["tipo_doc"] == "CPF":
            if tipo_doc == "CNPJ":
                student["tipo_doc"] = tipo_doc
        if vendedor:
            student["vendedor"] = vendedor
    else:
        planilha_students[email] = {
            "cpf": cpf,
            "nome": aluno["nome"],
            "email": email if "@noemail" not in email else "",
            "celular": celular,
            "endereco": endereco,
            "tipo_doc": tipo_doc,
            "vendas": [venda_detail] if venda_detail else [],
            "data_inicio": str(aluno["inicio"]) if aluno["inicio"] else (str(data_venda) if data_venda else ""),
            "data_fim": str(aluno["fim"]) if aluno["fim"] else "",
            "meses_curso": aluno["meses"],
            "cancelou": cancelamento,
            "data_cancelamento": str(data_cancelamento) if data_cancelamento else "",
            "renovacoes": 1 if renovacao_str in ("Renovação", "Sim") else 0,
            "total_gasto": valor_total,
            "produto_principal": produto,
            "nivel": nivel,
            "duracao": duracao_curso or str(aluno["meses"] or ""),
            "vendedor": vendedor,
            # Vindi placeholders
            "vindi_bills": 0,
            "vindi_total": 0.0,
            "vindi_last_payment": "",
            "vindi_first_payment": "",
            "vindi_bill_detail": [],
            # Source
            "fonte_planilha": True,
            "fonte_vindi": False,
        }

print(f"  Linked Alunos→Vendas: {linked}")
print(f"  Unlinked (no matching sale_key): {unlinked}")
print(f"  Unique planilha students: {len(planilha_students)}")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 4: Read Vindi bills (receitas_all_data.js) — 7,642 bills
# ═══════════════════════════════════════════════════════════════════════════════

print("\nSTEP 4: Reading Vindi bills from receitas_all_data.js...")

with open(os.path.join(PROJECT, "receitas_all_data.js"), "r") as f:
    raw_js = f.read()

# Parse: remove 'var data = ' prefix and trailing ';'
json_str = raw_js.replace("var data = ", "", 1).rstrip().rstrip(";")
vindi_bills_raw = json.loads(json_str)

print(f"  Total Vindi bills: {len(vindi_bills_raw)}")

# Group by cliente
vindi_by_client = defaultdict(list)
for bill in vindi_bills_raw:
    cliente = (bill.get("cliente", "") or "").strip()
    if cliente:
        vindi_by_client[cliente].append(bill)

print(f"  Unique Vindi clients: {len(vindi_by_client)}")

# Build aggregated Vindi records per client
vindi_aggregated = {}
for client_name, bills in vindi_by_client.items():
    total = sum(safe_float(b.get("valor", 0)) for b in bills)
    dates = [safe_date(b.get("data")) for b in bills]
    dates = [d for d in dates if d is not None]
    first_pay = min(dates) if dates else None
    last_pay = max(dates) if dates else None

    vindi_aggregated[client_name] = {
        "nome": client_name,
        "bills": len(bills),
        "total": total,
        "first": str(first_pay) if first_pay else "",
        "last": str(last_pay) if last_pay else "",
        "bill_detail": [
            {
                "data": b.get("data", ""),
                "valor": safe_float(b.get("valor", 0)),
                "situacao": b.get("situacao", ""),
                "mes": b.get("mes", ""),
                "categoria": b.get("categoria", ""),
            }
            for b in sorted(bills, key=lambda x: x.get("data", ""))
        ],
    }

# Also read receitas_customers.js for CPF/email enrichment
print("  Reading receitas_customers.js for CPF/email enrichment...")
with open(os.path.join(PROJECT, "receitas_customers.js"), "r") as f:
    cust_js = f.read()

cd_prefix = "var customerData = "
cd_start = cust_js.index(cd_prefix) + len(cd_prefix)
cd_end = cust_js.index(";\nvar customerBills")
customer_data = json.loads(cust_js[cd_start:cd_end])

# Build lookup: vindi client name -> (cpf, email)
vindi_meta = {}
for cname, cdata in customer_data.items():
    vindi_meta[cname] = {
        "cpf": normalize_cpf(cdata.get("cpf", "")),
        "email": (cdata.get("email", "") or "").strip().lower(),
    }

print(f"  Customer metadata loaded: {len(vindi_meta)} entries")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 5: Match Vindi to planilha by email
# ═══════════════════════════════════════════════════════════════════════════════

print("\nSTEP 5: Matching Vindi → Planilha...")

# Build reverse lookups for planilha students
planilha_by_email = {}
planilha_by_cpf = {}
planilha_by_norm_name = {}

for email_key, student in planilha_students.items():
    real_email = student["email"]
    if real_email:
        planilha_by_email[real_email] = email_key
    cpf = student["cpf"]
    if cpf:
        planilha_by_cpf[cpf] = email_key
    norm = normalize_name_for_match(student["nome"])
    if norm:
        planilha_by_norm_name[norm] = email_key

matched_vindi_clients = set()
match_by_email = 0
match_by_cpf = 0
match_by_name = 0

for client_name, vindi_rec in vindi_aggregated.items():
    meta = vindi_meta.get(client_name, {})
    v_email = meta.get("email", "")
    v_cpf = meta.get("cpf", "")
    matched_key = None

    # Try email match first (Vindi cliente email from receitas_customers)
    if v_email and v_email in planilha_by_email:
        matched_key = planilha_by_email[v_email]
        match_by_email += 1

    # Try CPF match
    if not matched_key and v_cpf and v_cpf in planilha_by_cpf:
        matched_key = planilha_by_cpf[v_cpf]
        match_by_cpf += 1

    # Try fuzzy name match (normalize accents, lowercase)
    if not matched_key:
        norm_name = normalize_name_for_match(client_name)
        if norm_name and norm_name in planilha_by_norm_name:
            matched_key = planilha_by_norm_name[norm_name]
            match_by_name += 1

    if matched_key:
        student = planilha_students[matched_key]
        student["vindi_bills"] = vindi_rec["bills"]
        student["vindi_total"] = vindi_rec["total"]
        student["vindi_first_payment"] = vindi_rec["first"]
        student["vindi_last_payment"] = vindi_rec["last"]
        student["vindi_bill_detail"] = vindi_rec["bill_detail"]
        student["fonte_vindi"] = True
        matched_vindi_clients.add(client_name)

total_matched = len(matched_vindi_clients)
print(f"  Matched by email: {match_by_email}")
print(f"  Matched by CPF: {match_by_cpf}")
print(f"  Matched by name (fuzzy): {match_by_name}")
print(f"  Total matched: {total_matched}")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 6: Add Vindi-only customers (not in planilha)
# ═══════════════════════════════════════════════════════════════════════════════

print("\nSTEP 6: Adding Vindi-only customers...")

# Collect all existing emails and CPFs to avoid duplicates
existing_emails = set()
existing_cpfs = set()
existing_norm_names = set()
for student in planilha_students.values():
    if student["email"]:
        existing_emails.add(student["email"])
    if student["cpf"]:
        existing_cpfs.add(student["cpf"])
    n = normalize_name_for_match(student["nome"])
    if n:
        existing_norm_names.add(n)

vindi_only_students = {}
vindi_only_count = 0

for client_name, vindi_rec in vindi_aggregated.items():
    if client_name in matched_vindi_clients:
        continue

    meta = vindi_meta.get(client_name, {})
    v_email = meta.get("email", "")
    v_cpf = meta.get("cpf", "")

    # Double-check not already present
    if v_email and v_email in existing_emails:
        continue
    if v_cpf and v_cpf in existing_cpfs:
        continue
    norm = normalize_name_for_match(client_name)
    if norm and norm in existing_norm_names:
        continue

    # Determine tipo_doc
    tipo_doc = "CNPJ" if v_cpf and len(v_cpf) > 11 else "CPF"

    key = v_email if v_email else (v_cpf if v_cpf else f"vindi_{client_name.lower().replace(' ', '_')}")

    vindi_only_students[key] = {
        "cpf": v_cpf,
        "nome": title_name(client_name),
        "email": v_email,
        "celular": "",
        "endereco": "",
        "tipo_doc": tipo_doc,
        "vendas": [],
        "data_inicio": vindi_rec["first"],
        "data_fim": vindi_rec["last"],
        "meses_curso": 0,
        "cancelou": False,
        "data_cancelamento": "",
        "renovacoes": 0,
        "total_gasto": 0.0,
        "produto_principal": "",
        "nivel": "",
        "duracao": "",
        "vendedor": "",
        "vindi_bills": vindi_rec["bills"],
        "vindi_total": vindi_rec["total"],
        "vindi_first_payment": vindi_rec["first"],
        "vindi_last_payment": vindi_rec["last"],
        "vindi_bill_detail": vindi_rec["bill_detail"],
        "fonte_planilha": False,
        "fonte_vindi": True,
    }
    vindi_only_count += 1

    # Track to avoid dupes within Vindi-only
    if v_email:
        existing_emails.add(v_email)
    if v_cpf:
        existing_cpfs.add(v_cpf)
    if norm:
        existing_norm_names.add(norm)

print(f"  Vindi-only customers added: {vindi_only_count}")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 7: Merge all students and classify
# ═══════════════════════════════════════════════════════════════════════════════

print("\nSTEP 7: Merging and classifying all students...")

all_students = {}
all_students.update(planilha_students)
all_students.update(vindi_only_students)

for key, s in all_students.items():
    # Classify modalidade
    s["modalidade"] = classify_modalidade(s["produto_principal"], s["tipo_doc"])

    # Determine tipo_cliente
    s["tipo_cliente"] = "PJ" if s["tipo_doc"] == "CNPJ" else "PF"

    # Determine data_fim if missing
    data_fim = safe_date(s["data_fim"])

    # Determine status
    s["status"] = determine_status(
        s["cancelou"],
        data_fim,
        s["vindi_last_payment"],
        s["fonte_planilha"],
    )

    s["contrato_ativo"] = "Ativo" in s["status"]

    # Estimate remaining classes
    dur = 0
    try:
        dur = int(s["duracao"]) if s["duracao"] else 0
    except (ValueError, TypeError):
        dur = 12
    s["aulas_remanescentes"] = estimate_remaining_classes(
        s["data_inicio"], dur if dur else 12, s["status"]
    )

    # Meses restantes
    s["meses_restantes"] = compute_meses_restantes(s["data_fim"], s["status"])

    # Fonte label
    if s["fonte_planilha"] and s["fonte_vindi"]:
        s["fonte"] = "Planilha+Vindi"
    elif s["fonte_planilha"]:
        s["fonte"] = "Planilha"
    else:
        s["fonte"] = "Vindi"

total_all = len(all_students)
planilha_count = sum(1 for s in all_students.values() if s["fonte_planilha"])
vindi_only_final = sum(1 for s in all_students.values() if not s["fonte_planilha"])
matched_count = sum(1 for s in all_students.values() if s["fonte_planilha"] and s["fonte_vindi"])

# Status counts
ativos = sum(1 for s in all_students.values() if "Ativo" in s["status"])
cancelados = sum(1 for s in all_students.values() if s["status"] == "Cancelado")
expirados = sum(1 for s in all_students.values() if s["status"] == "Expirado")
inativos = sum(1 for s in all_students.values() if s["status"] == "Inativo")

# Modalidade counts
community = sum(1 for s in all_students.values() if s["modalidade"] == "Community")
community_flow = sum(1 for s in all_students.values() if s["modalidade"] == "Community Flow")
espanhol = sum(1 for s in all_students.values() if s["modalidade"] == "Espanhol")
private = sum(1 for s in all_students.values() if s["modalidade"] == "Private")
in_company = sum(1 for s in all_students.values() if s["modalidade"] == "In-Company")

# Financeiro
renovacoes_total = sum(s["renovacoes"] for s in all_students.values())
total_planilha_valor = sum(s["total_gasto"] for s in all_students.values())
total_vindi_valor = sum(s["vindi_total"] for s in all_students.values())

print(f"\n{'=' * 70}")
print(f"CONSOLIDATED DATABASE")
print(f"{'=' * 70}")
print(f"  Total students: {total_all}")
print(f"  Planilha: {planilha_count} | Vindi-only: {vindi_only_final} | Matched: {matched_count}")
print(f"  Ativos: {ativos} | Cancelados: {cancelados} | Expirados: {expirados} | Inativos: {inativos}")
print(f"  Community: {community} | Flow: {community_flow} | Espanhol: {espanhol} | Private: {private} | In-Company: {in_company}")
print(f"  Renovacoes: {renovacoes_total}")
print(f"  Total planilha: R$ {total_planilha_valor:,.2f}")
print(f"  Total vindi: R$ {total_vindi_valor:,.2f}")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 8: Generate alunos_data.js
# ═══════════════════════════════════════════════════════════════════════════════

print(f"\nSTEP 8: Generating alunos_data.js...")


def serialize_student(s):
    """Convert student dict to JSON-safe dict matching HTML expectations."""
    return {
        "cpf": s["cpf"],
        "nome": s["nome"],
        "email": s["email"],
        "celular": s["celular"],
        "endereco": s.get("endereco", ""),
        "tipo_doc": s["tipo_doc"],
        "tipo_cliente": s["tipo_cliente"],
        "modalidade": s["modalidade"],
        "nivel": s["nivel"],
        "produto_principal": s["produto_principal"],
        "data_inicio": s["data_inicio"],
        "data_fim": s["data_fim"],
        "contrato_ativo": s["contrato_ativo"],
        "cancelou": s["cancelou"],
        "data_cancelamento": s["data_cancelamento"],
        "renovacoes": s["renovacoes"],
        "total_gasto": round(s["total_gasto"], 2),
        "status": s["status"],
        "aulas_remanescentes": s["aulas_remanescentes"],
        "meses_restantes": s.get("meses_restantes", 0),
        "vendas": s["vendas"],
        "vindi_bills": s["vindi_bills"],
        "vindi_total": round(s["vindi_total"], 2),
        "vindi_first_payment": s.get("vindi_first_payment", ""),
        "vindi_last_payment": s["vindi_last_payment"],
        "vindi_bill_detail": s["vindi_bill_detail"],
        "fonte_planilha": s["fonte_planilha"],
        "fonte_vindi": s["fonte_vindi"],
        "fonte": s["fonte"],
        "vendedor": s.get("vendedor", ""),
    }


sorted_students = sorted(all_students.values(), key=lambda s: s["nome"].lower())
alunos_array = [serialize_student(s) for s in sorted_students]

stats_obj = {
    "total": total_all,
    "ativos": ativos,
    "cancelados": cancelados,
    "expirados": expirados,
    "inativos": inativos,
    "planilha": planilha_count,
    "vindi_only": vindi_only_final,
    "matched": matched_count,
    "community": community,
    "community_flow": community_flow,
    "espanhol": espanhol,
    "private": private,
    "in_company": in_company,
    "renovacoes": renovacoes_total,
    "total_planilha": round(total_planilha_valor, 2),
    "total_vindi": round(total_vindi_valor, 2),
}

js_content = "var alunosData = " + json.dumps(alunos_array, ensure_ascii=False, indent=1) + ";\n\n"
js_content += "var alunosStats = " + json.dumps(stats_obj, ensure_ascii=False, indent=1) + ";\n"

with open(os.path.join(PROJECT, "alunos_data.js"), "w") as f:
    f.write(js_content)

print(f"  alunos_data.js written ({len(alunos_array)} students, {len(js_content):,} chars)")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 9: Generate Alunos_Base_Completa.xlsx (6 tabs)
# ═══════════════════════════════════════════════════════════════════════════════

print(f"\nSTEP 9: Generating Alunos_Base_Completa.xlsx...")

# Style definitions
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill_green = PatternFill(start_color="0D7C3F", end_color="0D7C3F", fill_type="solid")
header_fill_blue = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
header_fill_red = PatternFill(start_color="7A1F1F", end_color="7A1F1F", fill_type="solid")
header_fill_purple = PatternFill(start_color="4A1D8A", end_color="4A1D8A", fill_type="solid")
header_fill_cyan = PatternFill(start_color="0E7490", end_color="0E7490", fill_type="solid")
header_fill_amber = PatternFill(start_color="92400E", end_color="92400E", fill_type="solid")

fill_active = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
fill_cancel = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
fill_expired = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
fill_inactive = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

align_center = Alignment(horizontal="center", vertical="center")

XL_HEADERS = [
    "Nome", "Email", "CPF/CNPJ", "Celular", "Endereco", "Tipo Doc",
    "Modalidade", "Nivel", "Produto", "Data Inicio", "Data Fim",
    "Meses Curso", "Valor Planilha (R$)", "Renovacao", "Vendedor",
    "Cancelamento", "Data Cancelamento",
    "Status", "Vindi Bills", "Vindi Total (R$)",
    "Vindi Primeiro Pgto", "Vindi Ultimo Pgto", "Fonte",
]


def write_student_row(ws, row, s):
    """Write student row to worksheet."""
    vendedor = s.get("vendedor", "")
    if not vendedor and s["vendas"]:
        vendedor = s["vendas"][-1].get("vendedor", "")

    renovacao_label = ""
    if s["renovacoes"] > 0:
        renovacao_label = "Renovacao"
    elif s["vendas"]:
        last_renov = s["vendas"][-1].get("renovacao", "")
        if last_renov:
            renovacao_label = last_renov
    else:
        renovacao_label = "Novo"

    values = [
        s["nome"],
        s["email"],
        s["cpf"],
        s["celular"],
        s.get("endereco", ""),
        s["tipo_doc"],
        s["modalidade"],
        s["nivel"],
        s["produto_principal"],
        s["data_inicio"],
        s["data_fim"],
        s.get("meses_curso", 0) or "",
        round(s["total_gasto"], 2),
        renovacao_label,
        vendedor,
        "Sim" if s["cancelou"] else "",
        s["data_cancelamento"],
        s["status"],
        s["vindi_bills"],
        round(s["vindi_total"], 2),
        s.get("vindi_first_payment", ""),
        s["vindi_last_payment"],
        s.get("fonte", ""),
    ]

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
        if col in (10, 11, 17, 21, 22):  # date columns
            cell.alignment = align_center
        if col in (12, 13, 19, 20):  # number columns
            cell.alignment = align_center
            if col in (13, 20):
                cell.number_format = '#,##0.00'

    # Color status cell
    status_cell = ws.cell(row=row, column=18)
    status_val = s["status"]
    if "Ativo" in status_val:
        status_cell.fill = fill_active
    elif status_val == "Cancelado":
        status_cell.fill = fill_cancel
    elif status_val == "Expirado":
        status_cell.fill = fill_expired
    elif status_val == "Inativo":
        status_cell.fill = fill_inactive


def write_headers(ws, headers, fill):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = align_center
        cell.border = thin_border


def auto_width(ws, headers):
    for col, h in enumerate(headers, 1):
        max_len = len(h) + 2
        sample_rows = min(ws.max_row, 200)
        for row in range(2, sample_rows + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, min(len(str(val)) + 2, 50))
        ws.column_dimensions[get_column_letter(col)].width = max_len


def add_filters(ws, headers):
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


# Sort all by name
all_sorted = sorted(all_students.values(), key=lambda s: s["nome"].lower())

xwb = openpyxl.Workbook()

# Tab 1: Base Completa — ALL students
ws1 = xwb.active
ws1.title = "Base Completa"
write_headers(ws1, XL_HEADERS, header_fill_green)
for i, s in enumerate(all_sorted):
    write_student_row(ws1, i + 2, s)
auto_width(ws1, XL_HEADERS)
add_filters(ws1, XL_HEADERS)
ws1.freeze_panes = "A2"
print(f"  Base Completa: {len(all_sorted)} rows")

# Tab 2: Ativos — status contains "Ativo"
ws2 = xwb.create_sheet("Ativos")
ativos_list = [s for s in all_sorted if "Ativo" in s["status"]]
write_headers(ws2, XL_HEADERS, header_fill_blue)
for i, s in enumerate(ativos_list):
    write_student_row(ws2, i + 2, s)
auto_width(ws2, XL_HEADERS)
add_filters(ws2, XL_HEADERS)
ws2.freeze_panes = "A2"
print(f"  Ativos: {len(ativos_list)} rows")

# Tab 3: Cancelados
ws3 = xwb.create_sheet("Cancelados")
cancel_list = [s for s in all_sorted if s["status"] == "Cancelado"]
write_headers(ws3, XL_HEADERS, header_fill_red)
for i, s in enumerate(cancel_list):
    write_student_row(ws3, i + 2, s)
auto_width(ws3, XL_HEADERS)
add_filters(ws3, XL_HEADERS)
ws3.freeze_panes = "A2"
print(f"  Cancelados: {len(cancel_list)} rows")

# Tab 4: Renovacoes
ws4 = xwb.create_sheet(u"Renovações")
renov_list = [s for s in all_sorted if s["renovacoes"] > 0 or
              any(v.get("renovacao", "") in ("Renovação", "Sim") for v in s.get("vendas", []))]
write_headers(ws4, XL_HEADERS, header_fill_purple)
for i, s in enumerate(renov_list):
    write_student_row(ws4, i + 2, s)
auto_width(ws4, XL_HEADERS)
add_filters(ws4, XL_HEADERS)
ws4.freeze_panes = "A2"
print(f"  Renovacoes: {len(renov_list)} rows")

# Tab 5: Para Importacao Portal
ws5 = xwb.create_sheet(u"Para Importação Portal")
PORTAL_HEADERS = ["Nome", "Email", "CPF", "Celular", "Modalidade", "Nivel", "Status", "Meses Restantes"]
portal_list = [s for s in all_sorted if "Ativo" in s["status"]]
write_headers(ws5, PORTAL_HEADERS, header_fill_cyan)
for i, s in enumerate(portal_list):
    row = i + 2
    values = [
        s["nome"], s["email"], s["cpf"], s["celular"],
        s["modalidade"], s["nivel"], s["status"],
        s.get("meses_restantes", 0),
    ]
    for col, val in enumerate(values, 1):
        cell = ws5.cell(row=row, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
auto_width(ws5, PORTAL_HEADERS)
add_filters(ws5, PORTAL_HEADERS)
ws5.freeze_panes = "A2"
print(f"  Para Importacao Portal: {len(portal_list)} rows")

# Tab 6: Resumo
ws6 = xwb.create_sheet("Resumo")

summary_data = [
    ("Metrica", "Valor"),
    ("Total Alunos (Base Consolidada)", total_all),
    ("Planilha de Vendas", planilha_count),
    ("Apenas Vindi", vindi_only_final),
    ("Cruzados (Planilha + Vindi)", matched_count),
    ("", ""),
    ("STATUS", ""),
    ("Ativos", ativos),
    ("Cancelados", cancelados),
    ("Expirados", expirados),
    ("Inativos", inativos),
    ("", ""),
    ("MODALIDADE", ""),
    ("Community", community),
    ("Community Flow", community_flow),
    ("Espanhol", espanhol),
    ("Private", private),
    ("In-Company", in_company),
    ("", ""),
    ("FINANCEIRO", ""),
    ("Total Planilha (R$)", f"R$ {total_planilha_valor:,.2f}"),
    ("Total Vindi (R$)", f"R$ {total_vindi_valor:,.2f}"),
    ("Renovacoes", renovacoes_total),
    ("", ""),
    ("Gerado em", str(TODAY)),
]

for r, (metric, val) in enumerate(summary_data, 1):
    c1 = ws6.cell(row=r, column=1, value=metric)
    c2 = ws6.cell(row=r, column=2, value=val)
    c1.border = thin_border
    c2.border = thin_border
    if r == 1:
        c1.font = header_font
        c1.fill = header_fill_amber
        c2.font = header_font
        c2.fill = header_fill_amber
    elif metric in ("STATUS", "MODALIDADE", "FINANCEIRO"):
        c1.font = Font(bold=True, size=11)

ws6.column_dimensions["A"].width = 40
ws6.column_dimensions["B"].width = 30

# Save Excel
excel_path = os.path.join(PROJECT, "Alunos_Base_Completa.xlsx")
xwb.save(excel_path)
print(f"  Excel saved: {excel_path}")

# ═══════════════════════════════════════════════════════════════════════════════
# DONE
# ═══════════════════════════════════════════════════════════════════════════════

print(f"\n{'=' * 70}")
print("BUILD COMPLETE")
print(f"{'=' * 70}")
print(f"  alunos_data.js: {len(alunos_array)} students")
print(f"  Alunos_Base_Completa.xlsx: 6 tabs")
print(f"  Total: {total_all} | Planilha: {planilha_count} | Vindi-only: {vindi_only_final} | Matched: {matched_count}")
print(f"  Ativos: {ativos} | Cancelados: {cancelados} | Expirados: {expirados} | Inativos: {inativos}")
